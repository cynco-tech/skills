#!/usr/bin/env python3
"""
[FIRM_NAME] — Standardised PDF Financial Statements Generator

Usage:
    python generate_pdf_report.py <input_xlsx> <output_pdf> [--entity-type sdn_bhd|sole_prop|partnership]

Reads from the standard Excel working papers workbook and produces
a clean, minimal, black-and-white PDF report.

Requirements:
    pip install reportlab openpyxl
"""

import argparse
import sys
import openpyxl
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, KeepTogether
)
from reportlab.lib import colors

# ══════════════════════════════════════════════════════════════
# CONSTANTS
# ══════════════════════════════════════════════════════════════
PAGE_W, PAGE_H = A4
MARGIN_L = 20 * mm
MARGIN_R = 20 * mm
MARGIN_T = 18 * mm
MARGIN_B = 18 * mm
CONTENT_W = PAGE_W - MARGIN_L - MARGIN_R

FIRM = "[FIRM_NAME]"
FIRM_SUB = "Chartered Accountants ([CA_REGISTRATION])"

BLACK = colors.Color(0, 0, 0)
DARK_GREY = colors.Color(0.2, 0.2, 0.2)
MID_GREY = colors.Color(0.45, 0.45, 0.45)
LIGHT_RULE = colors.Color(0.7, 0.7, 0.7)
VERY_LIGHT = colors.Color(0.92, 0.92, 0.92)
ALT_ROW = colors.Color(0.97, 0.97, 0.97)

# ══════════════════════════════════════════════════════════════
# STYLES
# ══════════════════════════════════════════════════════════════
# Cover
sty_cover_client = ParagraphStyle('CC', fontName='Helvetica-Bold', fontSize=22,
    leading=28, alignment=TA_CENTER, textColor=BLACK, spaceAfter=6*mm)
sty_cover_reg = ParagraphStyle('CR', fontName='Helvetica', fontSize=10,
    leading=14, alignment=TA_CENTER, textColor=MID_GREY, spaceAfter=18*mm)
sty_cover_title = ParagraphStyle('CT', fontName='Helvetica', fontSize=13,
    leading=18, alignment=TA_CENTER, textColor=DARK_GREY, spaceAfter=2*mm)
sty_cover_fy = ParagraphStyle('CF', fontName='Helvetica', fontSize=11,
    leading=15, alignment=TA_CENTER, textColor=MID_GREY, spaceAfter=30*mm)
sty_cover_firm = ParagraphStyle('CFi', fontName='Helvetica-Bold', fontSize=11,
    leading=15, alignment=TA_CENTER, textColor=BLACK, spaceAfter=1*mm)
sty_cover_firm_sub = ParagraphStyle('CFs', fontName='Helvetica', fontSize=9,
    leading=13, alignment=TA_CENTER, textColor=MID_GREY)

# Headers
sty_h1 = ParagraphStyle('H1', fontName='Helvetica-Bold', fontSize=13, leading=17,
    alignment=TA_CENTER, textColor=BLACK, spaceAfter=1*mm)
sty_h2 = ParagraphStyle('H2', fontName='Helvetica', fontSize=10, leading=14,
    alignment=TA_CENTER, textColor=MID_GREY, spaceAfter=1*mm)
sty_h3 = ParagraphStyle('H3', fontName='Helvetica', fontSize=9, leading=13,
    alignment=TA_CENTER, textColor=MID_GREY, spaceAfter=6*mm)

# Body
sty_section = ParagraphStyle('Sec', fontName='Helvetica-Bold', fontSize=9,
    leading=13, textColor=BLACK, spaceBefore=4*mm, spaceAfter=1*mm)
sty_normal = ParagraphStyle('N', fontName='Helvetica', fontSize=8.5,
    leading=11, textColor=DARK_GREY)
sty_normal_r = ParagraphStyle('NR', fontName='Helvetica', fontSize=8.5,
    leading=11, textColor=DARK_GREY, alignment=TA_RIGHT)
sty_bold = ParagraphStyle('B', fontName='Helvetica-Bold', fontSize=8.5,
    leading=11, textColor=BLACK)
sty_bold_r = ParagraphStyle('BR', fontName='Helvetica-Bold', fontSize=8.5,
    leading=11, textColor=BLACK, alignment=TA_RIGHT)
sty_small = ParagraphStyle('S', fontName='Helvetica', fontSize=7.5,
    leading=10, textColor=MID_GREY)
sty_small_r = ParagraphStyle('SR', fontName='Helvetica', fontSize=7.5,
    leading=10, textColor=MID_GREY, alignment=TA_RIGHT)
sty_gl_acct = ParagraphStyle('GLA', fontName='Helvetica-Bold', fontSize=8.5,
    leading=12, textColor=BLACK, spaceBefore=3*mm, spaceAfter=1*mm)
sty_note = ParagraphStyle('Note', fontName='Helvetica-Oblique', fontSize=7.5,
    leading=10, textColor=MID_GREY, alignment=TA_CENTER)

# ══════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════
def fmt(v, dash_zero=False):
    if v is None or v == 0:
        return '-'
    if isinstance(v, str):
        return v
    neg = v < 0
    s = f"{abs(v):,.2f}"
    return f"({s})" if neg else s

def make_header(client_name, title, subtitle):
    return [
        Paragraph(client_name, sty_h1),
        Paragraph(title, sty_h2),
        Paragraph(subtitle, sty_h3),
    ]

def thin_line(w):
    t = Table([['']],colWidths=[w],rowHeights=[0.3*mm])
    t.setStyle(TableStyle([('LINEBELOW',(0,0),(-1,-1),0.5,LIGHT_RULE)]))
    return t

# ══════════════════════════════════════════════════════════════
# PAGE CALLBACKS
# ══════════════════════════════════════════════════════════════
page_count = [0]

def on_page(canvas_obj, doc):
    page_count[0] += 1
    pn = page_count[0]
    if pn > 1:
        canvas_obj.saveState()
        canvas_obj.setFont('Helvetica', 7)
        canvas_obj.setFillColor(MID_GREY)
        canvas_obj.drawString(MARGIN_L, 10*mm, f"{FIRM} {FIRM_SUB}")
        canvas_obj.drawRightString(PAGE_W - MARGIN_R, 10*mm, f"Page {pn - 1}")
        canvas_obj.setStrokeColor(LIGHT_RULE)
        canvas_obj.setLineWidth(0.4)
        canvas_obj.line(MARGIN_L, 13*mm, PAGE_W - MARGIN_R, 13*mm)
        canvas_obj.restoreState()

def on_first_page(canvas_obj, doc):
    page_count[0] += 1

# ══════════════════════════════════════════════════════════════
# DATA EXTRACTION FROM XLSX
# ══════════════════════════════════════════════════════════════
def extract_data(xlsx_path):
    """Extract all financial data from the standard working papers workbook."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    data = {}

    # Company Info
    if 'Company Info' in wb.sheetnames:
        ws = wb['Company Info']
        info = {}
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            if row[0] and row[1]:
                info[str(row[0]).strip()] = str(row[1]).strip()
            elif row[0] and not row[1]:
                info['_name'] = str(row[0]).strip()
        data['company_info'] = info

    # Income Statement
    is_data = []
    if 'Income Statement' in wb.sheetnames:
        ws = wb['Income Statement']
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
            if row[0] is not None:
                is_data.append((str(row[0]), row[1]))
    data['income_statement'] = is_data

    # Balance Sheet
    bs_data = []
    if 'Balance Sheet' in wb.sheetnames:
        ws = wb['Balance Sheet']
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
            if row[0] is not None:
                bs_data.append((str(row[0]), row[1]))
    data['balance_sheet'] = bs_data

    # Trial Balance
    tb_data = []
    if 'Trial Balance' in wb.sheetnames:
        ws = wb['Trial Balance']
        for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=True):
            if row[0] is not None or (row[1] is not None and row[1] not in (None, '')):
                tb_data.append((row[0], row[1], row[2], row[3]))
    data['trial_balance'] = tb_data

    # General Ledger
    gl_accounts = []
    if 'General Ledger' in wb.sheetnames:
        ws = wb['General Ledger']
        current_acct = None
        current_rows = []
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            r0 = row[0]
            if r0 and isinstance(r0, str) and ' - ' in r0 and row[1] is None and row[3] is None:
                if current_acct:
                    gl_accounts.append((current_acct, current_rows))
                current_acct = r0
                current_rows = []
            elif r0 == 'Date':
                continue
            elif current_acct and (r0 is not None or (row[2] is not None and row[2] != '')):
                current_rows.append(row)
        if current_acct:
            gl_accounts.append((current_acct, current_rows))
    data['general_ledger'] = gl_accounts

    # Queries (if sheet exists)
    queries = []
    for sheet_name in ['Queries & Notes', 'Queries', 'Notes']:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                if any(c is not None for c in row):
                    queries.append(row)
            break
    data['queries'] = queries

    return data

# ══════════════════════════════════════════════════════════════
# PDF BUILDERS
# ══════════════════════════════════════════════════════════════
def build_cover(story, client_name, reg_no, fy_desc):
    story.append(Spacer(1, 55*mm))
    story.append(thin_line(CONTENT_W))
    story.append(Spacer(1, 12*mm))
    story.append(Paragraph(client_name, sty_cover_client))
    if reg_no:
        story.append(Paragraph(f"({reg_no})", sty_cover_reg))
    else:
        story.append(Spacer(1, 18*mm))
    story.append(Paragraph("FINANCIAL STATEMENTS", sty_cover_title))
    story.append(Paragraph(fy_desc.upper(), sty_cover_fy))
    story.append(Spacer(1, 5*mm))
    story.append(thin_line(CONTENT_W))
    story.append(Spacer(1, 12*mm))
    story.append(Paragraph("Prepared by", ParagraphStyle('PB', fontName='Helvetica',
        fontSize=9, leading=13, alignment=TA_CENTER, textColor=MID_GREY, spaceAfter=3*mm)))
    story.append(Paragraph(FIRM, sty_cover_firm))
    story.append(Paragraph(FIRM_SUB, sty_cover_firm_sub))
    story.append(PageBreak())

def build_toc(story, client_name, sections):
    story.extend(make_header(client_name, "TABLE OF CONTENTS", " "))
    story.append(Spacer(1, 8*mm))
    col_w = [CONTENT_W - 20*mm, 20*mm]
    rows = []
    for title, pg in sections:
        rows.append([
            Paragraph(title, ParagraphStyle('TI', fontName='Helvetica', fontSize=10,
                leading=18, textColor=DARK_GREY)),
            Paragraph(str(pg), ParagraphStyle('TP', fontName='Helvetica', fontSize=10,
                leading=18, textColor=DARK_GREY, alignment=TA_RIGHT)),
        ])
    t = Table(rows, colWidths=col_w)
    t.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LINEBELOW', (0,0), (-1,-1), 0.3, LIGHT_RULE),
        ('TOPPADDING', (0,0), (-1,-1), 4*mm),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4*mm),
        ('LEFTPADDING', (0,0), (0,-1), 2*mm),
        ('RIGHTPADDING', (-1,0), (-1,-1), 2*mm),
    ]))
    story.append(t)
    story.append(PageBreak())

def build_two_col_statement(story, client_name, title, subtitle, data, skip_labels=None):
    """Build a P&L or BS style two-column statement."""
    story.extend(make_header(client_name, title, subtitle))
    col_w = [CONTENT_W * 0.65, CONTENT_W * 0.35]
    rows = []
    style_cmds = [
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING', (0,0), (0,-1), 2*mm),
        ('RIGHTPADDING', (-1,0), (-1,-1), 2*mm),
        ('TOPPADDING', (0,0), (-1,-1), 1.2*mm),
        ('BOTTOMPADDING', (0,0), (-1,-1), 1.2*mm),
    ]
    skip = skip_labels or []

    for i, (label, val) in enumerate(data):
        if label in skip:
            continue
        is_section = any(label == s for s in ['REVENUE', 'LESS: OPERATING EXPENSES',
            'ASSETS', 'LIABILITIES AND EQUITY', 'Current Assets', 'Non-Current Assets',
            'Current Liabilities', 'Equity'])
        is_total = label.startswith('TOTAL') or label.startswith('Total')
        is_profit = label.startswith('PROFIT') or label.startswith('NET')
        is_indented = label.startswith('  ')

        if is_section:
            if label in ('REVENUE', 'ASSETS'):
                rows.append([Paragraph(f"<b>{label}</b>", sty_bold),
                    Paragraph('<b>RM</b>', sty_bold_r)])
                style_cmds.append(('LINEBELOW', (1, len(rows)-1), (1, len(rows)-1), 0.5, BLACK))
            elif label in ('Current Assets', 'Non-Current Assets', 'Current Liabilities', 'Equity'):
                rows.append([Paragraph(f"<b>{label}</b>", sty_section), Paragraph('', sty_normal)])
            else:
                rows.append([Paragraph(f"<b>{label}</b>", sty_bold), Paragraph('', sty_normal)])
        elif is_profit:
            rows.append([Paragraph(f"<b>{label}</b>", sty_bold),
                Paragraph(f"<b>{fmt(val)}</b>", sty_bold_r)])
            style_cmds.append(('LINEABOVE', (0, len(rows)-1), (-1, len(rows)-1), 0.8, BLACK))
            style_cmds.append(('LINEBELOW', (1, len(rows)-1), (1, len(rows)-1), 1.0, BLACK))
        elif is_total:
            rows.append([Paragraph(f"<b>{label}</b>", sty_bold),
                Paragraph(f"<b>{fmt(val)}</b>", sty_bold_r)])
            if label.startswith('TOTAL'):
                style_cmds.append(('LINEABOVE', (0, len(rows)-1), (-1, len(rows)-1), 0.8, BLACK))
                style_cmds.append(('LINEBELOW', (1, len(rows)-1), (1, len(rows)-1), 1.0, BLACK))
            else:
                style_cmds.append(('LINEABOVE', (1, len(rows)-1), (1, len(rows)-1), 0.5, BLACK))
                style_cmds.append(('LINEBELOW', (1, len(rows)-1), (1, len(rows)-1), 0.5, BLACK))
        elif is_indented:
            rows.append([Paragraph(f"&nbsp;&nbsp;&nbsp;&nbsp;{label.strip()}", sty_normal),
                Paragraph(fmt(val, dash_zero=True), sty_normal_r)])
        else:
            rows.append([Paragraph(label, sty_normal),
                Paragraph(fmt(val, dash_zero=True), sty_normal_r)])

    t = Table(rows, colWidths=col_w)
    t.setStyle(TableStyle(style_cmds))
    story.append(t)
    story.append(Spacer(1, 15*mm))
    story.append(Paragraph("The accompanying notes form an integral part of these financial statements.", sty_note))
    story.append(PageBreak())

def build_trial_balance(story, client_name, subtitle, tb_data):
    story.extend(make_header(client_name, "TRIAL BALANCE", subtitle))
    col_w = [15*mm, CONTENT_W - 15*mm - 30*mm - 30*mm, 30*mm, 30*mm]
    rows = [[
        Paragraph("<b>Code</b>", sty_bold),
        Paragraph("<b>Account Name</b>", sty_bold),
        Paragraph("<b>Debit (RM)</b>", sty_bold_r),
        Paragraph("<b>Credit (RM)</b>", sty_bold_r),
    ]]
    style_cmds = [
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING', (0,0), (-1,-1), 1.5*mm),
        ('RIGHTPADDING', (0,0), (-1,-1), 1.5*mm),
        ('TOPPADDING', (0,0), (-1,-1), 1.0*mm),
        ('BOTTOMPADDING', (0,0), (-1,-1), 1.0*mm),
        ('LINEBELOW', (0,0), (-1,0), 0.8, BLACK),
        ('BACKGROUND', (0,0), (-1,0), VERY_LIGHT),
    ]

    for code, name, dr, cr in tb_data:
        if code is None and name == 'TOTAL':
            rows.append([Paragraph('', sty_normal), Paragraph('<b>TOTAL</b>', sty_bold),
                Paragraph(f"<b>{fmt(dr)}</b>", sty_bold_r),
                Paragraph(f"<b>{fmt(cr)}</b>", sty_bold_r)])
        elif code is None and name == 'Difference':
            rows.append([Paragraph('', sty_normal), Paragraph('<b>Difference</b>', sty_bold),
                Paragraph(f"<b>{fmt(dr)}</b>", sty_bold_r), Paragraph('', sty_normal)])
        elif code is not None:
            rows.append([Paragraph(str(code), sty_small), Paragraph(str(name or ''), sty_normal),
                Paragraph(fmt(dr, True), sty_normal_r), Paragraph(fmt(cr, True), sty_normal_r)])

    # Alternating rows
    for i in range(1, len(rows)):
        if i % 2 == 0:
            style_cmds.append(('BACKGROUND', (0,i), (-1,i), ALT_ROW))

    # Line above TOTAL
    total_idx = len(rows) - 2
    style_cmds.append(('LINEABOVE', (0, total_idx), (-1, total_idx), 0.8, BLACK))
    style_cmds.append(('LINEBELOW', (2, total_idx), (3, total_idx), 1.0, BLACK))

    t = Table(rows, colWidths=col_w)
    t.setStyle(TableStyle(style_cmds))
    story.append(t)
    story.append(PageBreak())

def build_general_ledger(story, client_name, subtitle, gl_accounts):
    story.extend(make_header(client_name, "GENERAL LEDGER", subtitle))
    story.append(Spacer(1, 2*mm))
    col_w = [18*mm, 10*mm, CONTENT_W - 18*mm - 10*mm - 24*mm - 24*mm - 26*mm, 24*mm, 24*mm, 26*mm]

    for acct_name, rows_data in gl_accounts:
        block = []
        block.append(Paragraph(f"<b>{acct_name}</b>", sty_gl_acct))

        gl_rows = [[
            Paragraph("<b>Date</b>", sty_small),
            Paragraph("<b>JE</b>", sty_small),
            Paragraph("<b>Description</b>", sty_small),
            Paragraph("<b>Debit</b>", sty_small_r),
            Paragraph("<b>Credit</b>", sty_small_r),
            Paragraph("<b>Balance</b>", sty_small_r),
        ]]

        for r in rows_data:
            dt_val = r[0]
            if isinstance(dt_val, datetime):
                dt_str = dt_val.strftime('%d/%m/%Y')
            elif dt_val is None:
                dt_str = ''
            else:
                dt_str = str(dt_val)

            je_no = str(r[1]) if r[1] is not None else ''
            desc = str(r[2]) if r[2] else ''
            if len(desc) > 45:
                desc = desc[:43] + '..'
            dr, cr, bal = r[3], r[4], r[5]
            is_closing = (desc == 'Closing Balance')

            gl_rows.append([
                Paragraph(dt_str, sty_small if not is_closing else sty_bold),
                Paragraph(je_no, sty_small),
                Paragraph(f"<b>{desc}</b>" if is_closing else desc, sty_small if not is_closing else sty_bold),
                Paragraph(fmt(dr, True) if not is_closing else '', sty_small_r),
                Paragraph(fmt(cr, True) if not is_closing else '', sty_small_r),
                Paragraph(f"<b>{fmt(bal)}</b>" if is_closing else fmt(bal),
                    sty_bold_r if is_closing else sty_small_r),
            ])

        gl_style = [
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('LEFTPADDING', (0,0), (-1,-1), 1*mm),
            ('RIGHTPADDING', (0,0), (-1,-1), 1*mm),
            ('TOPPADDING', (0,0), (-1,-1), 0.6*mm),
            ('BOTTOMPADDING', (0,0), (-1,-1), 0.6*mm),
            ('LINEBELOW', (0,0), (-1,0), 0.5, BLACK),
            ('BACKGROUND', (0,0), (-1,0), VERY_LIGHT),
        ]
        if len(gl_rows) > 1:
            gl_style.append(('LINEABOVE', (0, len(gl_rows)-1), (-1, len(gl_rows)-1), 0.5, BLACK))

        t = Table(gl_rows, colWidths=col_w, repeatRows=1)
        t.setStyle(TableStyle(gl_style))
        block.append(t)
        block.append(Spacer(1, 3*mm))

        if len(rows_data) <= 15:
            story.append(KeepTogether(block))
        else:
            story.extend(block)

# ══════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(description='Generate PDF Financial Statements')
    parser.add_argument('input_xlsx', help='Path to Excel working papers')
    parser.add_argument('output_pdf', help='Output PDF path')
    parser.add_argument('--entity-type', default='sdn_bhd', choices=['sdn_bhd', 'sole_prop', 'partnership'])
    args = parser.parse_args()

    # Extract data
    data = extract_data(args.input_xlsx)
    info = data.get('company_info', {})

    client_name = info.get('_name', 'CLIENT NAME')
    reg_no = info.get('Company Registration No', '')
    fy_desc = "Financial Year Ended 31 December 2024"  # Default; agent should override

    # Try to detect FY from sheet headers
    if data['income_statement']:
        for label, _ in data['income_statement']:
            if 'YEAR' in str(label).upper() and 'ENDED' in str(label).upper():
                fy_desc = label
                break

    # Build PDF
    doc = SimpleDocTemplate(args.output_pdf, pagesize=A4,
        leftMargin=MARGIN_L, rightMargin=MARGIN_R,
        topMargin=MARGIN_T, bottomMargin=MARGIN_B)

    story = []

    # Determine TOC page numbers (approximate)
    toc_sections = [
        ("Statement of Comprehensive Income (Profit & Loss)", "3"),
        ("Statement of Financial Position (Balance Sheet)", "4"),
        ("Trial Balance", "5"),
        ("General Ledger", "6"),
    ]
    if data['queries']:
        toc_sections.append(("Queries & Outstanding Items", "..."))

    # Build
    build_cover(story, client_name, reg_no, fy_desc)
    build_toc(story, client_name, toc_sections)

    build_two_col_statement(story, client_name,
        "STATEMENT OF COMPREHENSIVE INCOME",
        "For the " + fy_desc,
        data['income_statement'])

    build_two_col_statement(story, client_name,
        "STATEMENT OF FINANCIAL POSITION",
        "As at " + fy_desc.replace("Financial Year Ended ", ""),
        data['balance_sheet'],
        skip_labels=['Balance Check (A - L&E)'])

    build_trial_balance(story, client_name,
        "As at " + fy_desc.replace("Financial Year Ended ", ""),
        data['trial_balance'])

    build_general_ledger(story, client_name,
        "For the " + fy_desc,
        data['general_ledger'])

    # Build
    page_count[0] = 0
    doc.build(story, onFirstPage=on_first_page, onLaterPages=on_page)
    print(f"PDF saved: {args.output_pdf}")
    print(f"Pages: {page_count[0]}")

if __name__ == '__main__':
    main()
