#!/usr/bin/env python3
"""
[CLIENT_NAME] — FY[YEAR] Working Papers Generator
FY: [FY_START] to [FY_END]
Framework: MPERS | Tax: ITA 1967

Generates a 12-sheet formula-driven Excel workbook.
Every total, subtotal, and cross-reference is an Excel formula.
"""

import csv, sys
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from collections import defaultdict, OrderedDict
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

BASE = Path(".")
OUTPUT = BASE / "[CLIENT]_FY[YEAR]_Working_Papers.xlsx"

D = Decimal


def d(value) -> Decimal:
    """Convert any numeric value to Decimal, rounding to 2dp."""
    return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


# ---------------------------------------------------------------------------
# RowTracker — tracks cell positions across sheets for cross-reference formulas
# ---------------------------------------------------------------------------

class RowTracker:
    """Tracks cell positions across sheets for cross-reference formulas."""

    def __init__(self):
        self._positions: dict[str, dict[str, int]] = defaultdict(dict)

    def record(self, sheet: str, key: str, row: int):
        self._positions[sheet][key] = row

    def row(self, sheet: str, key: str) -> int:
        return self._positions[sheet][key]

    def cell(self, sheet: str, col: str, key: str) -> str:
        return f"{col}{self._positions[sheet][key]}"

    def formula(self, sheet: str, col: str, key: str) -> str:
        r = self._positions[sheet][key]
        return f"='{sheet}'!{col}{r}"

    def has(self, sheet: str, key: str) -> bool:
        return key in self._positions.get(sheet, {})


# ---------------------------------------------------------------------------
# Chart of Accounts
# ---------------------------------------------------------------------------

COA = OrderedDict([
    # Assets
    ("1000", {"name": "Cash at Bank ([BANK_NAME] [ACCOUNT_NUMBER])", "type": "Asset", "normal": "DR"}),
    ("1010", {"name": "Cash in Hand", "type": "Asset", "normal": "DR"}),
    ("1100", {"name": "Trade Receivables", "type": "Asset", "normal": "DR"}),
    ("1210", {"name": "Prepayments", "type": "Asset", "normal": "DR"}),
    ("1220", {"name": "Deposits", "type": "Asset", "normal": "DR"}),
    ("1300", {"name": "PPE - Music System & Equipment", "type": "Asset", "normal": "DR"}),
    ("1310", {"name": "PPE - Computer & POS", "type": "Asset", "normal": "DR"}),
    ("1398", {"name": "Accumulated Depreciation - Computer & POS", "type": "Asset", "normal": "CR"}),
    ("1399", {"name": "Accumulated Depreciation - Music System & Equipment", "type": "Asset", "normal": "CR"}),
    # Liabilities
    ("2000", {"name": "Trade Payables", "type": "Liability", "normal": "CR"}),
    ("2100", {"name": "Other Payables & Accruals", "type": "Liability", "normal": "CR"}),
    ("2110", {"name": "EPF Payable", "type": "Liability", "normal": "CR"}),
    ("2120", {"name": "SOCSO Payable", "type": "Liability", "normal": "CR"}),
    ("2130", {"name": "EIS Payable", "type": "Liability", "normal": "CR"}),
    ("2350", {"name": "Amount Due to Director", "type": "Liability", "normal": "CR"}),
    ("2360", {"name": "Amount Due to Shareholders", "type": "Liability", "normal": "CR"}),
    ("2900", {"name": "Suspense Account", "type": "Liability", "normal": "CR"}),
    # Equity
    ("3000", {"name": "Share Capital", "type": "Equity", "normal": "CR"}),
    ("3200", {"name": "Retained Earnings (Prior Years)", "type": "Equity", "normal": "CR"}),
    # Revenue
    ("4000", {"name": "Event Revenue", "type": "Revenue", "normal": "CR"}),
    ("4100", {"name": "F&B / Cafe Revenue", "type": "Revenue", "normal": "CR"}),
    ("4200", {"name": "Other Income", "type": "Revenue", "normal": "CR"}),
    # COGS
    ("5050", {"name": "Food & Beverage Cost", "type": "Expense", "normal": "DR"}),
    # Expenses
    ("5000", {"name": "Salary & Wages", "type": "Expense", "normal": "DR"}),
    ("5030", {"name": "Performance & Show Expenses", "type": "Expense", "normal": "DR"}),
    ("5200", {"name": "Rental Expense", "type": "Expense", "normal": "DR"}),
    ("5210", {"name": "Utilities", "type": "Expense", "normal": "DR"}),
    ("5300", {"name": "Depreciation - Music Equipment", "type": "Expense", "normal": "DR"}),
    ("5310", {"name": "Depreciation - Computer & POS", "type": "Expense", "normal": "DR"}),
    ("5400", {"name": "Software & Subscriptions", "type": "Expense", "normal": "DR"}),
    ("5510", {"name": "Accounting Fee (Accrued)", "type": "Expense", "normal": "DR"}),
    ("5520", {"name": "Audit Fee (Accrued)", "type": "Expense", "normal": "DR"}),
    ("5950", {"name": "Printing & Stationery", "type": "Expense", "normal": "DR"}),
    ("5960", {"name": "Repairs & Maintenance", "type": "Expense", "normal": "DR"}),
    ("5970", {"name": "Miscellaneous Expenses", "type": "Expense", "normal": "DR"}),
    ("5980", {"name": "Bank Charges", "type": "Expense", "normal": "DR"}),
    # Tax
    ("9000", {"name": "Tax Expense", "type": "Expense", "normal": "DR"}),
])


# ---------------------------------------------------------------------------
# Classification Map — maps CSV account codes to GL account codes
# ---------------------------------------------------------------------------

CLASSIFICATION_MAP = {
    "4000": "4000",  "4100": "4100",  "4200": "4200",
    "5000": "5000",  "5030": "5030",
    "5100": "5050",  # F&B Cost (CSV) -> COGS (GL)
    "5200": "5200",  "5210": "5210",  "5400": "5400",
    "5950": "5950",  "5960": "5960",  "5970": "5970",  "5980": "5980",
    "2000": "2000",
    "2300": "2350",  # Due to Director (CSV) -> GL
    "2400": "2360",  # Due to Shareholders (CSV) -> GL
    "2100": "2100",
    "2900": "2100",  # Suspense -> Other Payables (unused but present per policy)
}


# ---------------------------------------------------------------------------
# Opening Balances (as at [FY_START])
# ---------------------------------------------------------------------------

OPENING_BALANCES = {
    "1000": d("15266.41"),
    "1010": d("6035.33"),
    "1100": d("17907.10"),
    "1210": d("333.33"),
    "1220": d("500.00"),
    "1300": d("60353.92"),
    "1310": d("3568.00"),
    "1398": d("-713.60"),
    "1399": d("-12070.78"),
    "2000": d("-64810.00"),
    "2100": d("-15715.00"),
    "2350": d("-17150.00"),
    "2360": d("-51052.06"),
    "3000": d("-1000.00"),
    "3200": d("58547.35"),   # Accumulated loss (DR balance)
}

# Verify opening balances tie
total_dr = sum(v for v in OPENING_BALANCES.values() if v > 0)
total_cr = sum(abs(v) for v in OPENING_BALANCES.values() if v < 0)
print(f"Opening DR: {total_dr:,.2f}")
print(f"Opening CR: {total_cr:,.2f}")
print(f"Difference: {total_dr - total_cr:,.2f}")
assert abs(total_dr - total_cr) < Decimal("0.01"), f"Opening balances don't tie! Diff={total_dr - total_cr:.2f}"
print("Opening balances TIE ✓")


# ---------------------------------------------------------------------------
# Prior Year Accrual Settlements
# ---------------------------------------------------------------------------

PRIOR_YEAR_SETTLEMENTS = [
    ("2024-02-08", "MUHAMMAD RIDUAN", "5000", d("-3945.00"), "2100"),
    ("2024-02-08", "MUHAMMAD AQMAR", "5000", d("-2370.00"), "2100"),
    ("2024-02-08", "VANDOLF CASTILLO", "5000", d("-1900.00"), "2100"),
    ("2024-02-08", "TENAGA NASIONAL", "5210", d("-3000.00"), "2100"),
]


# ---------------------------------------------------------------------------
# CSV Loading & Reclassification
# ---------------------------------------------------------------------------

def load_transactions() -> list[dict]:
    """Load and reclassify bank transactions."""
    all_txns = []
    for fname in ["classified_transactions.csv", "classified_jan2025.csv"]:
        fpath = BASE / fname
        if fpath.exists():
            with open(fpath) as f:
                all_txns.extend(list(csv.DictReader(f)))

    reclassified = 0
    for txn in all_txns:
        for settle_date, desc_match, orig_code, amt, new_code in PRIOR_YEAR_SETTLEMENTS:
            if (txn["date"] == settle_date
                    and desc_match in txn["description"].upper()
                    and txn["acct_code"] == orig_code
                    and d(txn["amount"]) == amt):
                txn["acct_code"] = new_code
                txn["acct_name"] = "Settlement of prior year accrual"
                reclassified += 1
                break

    print(f"Loaded {len(all_txns)} transactions, reclassified {reclassified} prior year settlements")
    if reclassified != 4:
        print(f"WARNING: Expected 4 reclassifications, got {reclassified}. Check PRIOR_YEAR_SETTLEMENTS matching.")
    return all_txns


# === MAIN ===
if __name__ == "__main__":
    all_txns = load_transactions()
    print(f"\nCOA accounts: {len(COA)}")
    print(f"Transactions: {len(all_txns)}")

    # ===================================================================
    # 1. JOURNAL ENTRY GENERATION
    # ===================================================================

    journal_entries: list[dict] = []
    je_num = 0

    # --- JE-001: Opening Balances ---
    je_num += 1
    for acct, bal in OPENING_BALANCES.items():
        if bal == d("0"):
            continue
        acct_info = COA[acct]
        if bal > 0:
            journal_entries.append({
                "je_num": je_num, "date": "2024-02-01",
                "acct": acct, "acct_name": acct_info["name"],
                "description": "Opening balance",
                "dr": bal, "cr": d("0"),
            })
        else:
            journal_entries.append({
                "je_num": je_num, "date": "2024-02-01",
                "acct": acct, "acct_name": acct_info["name"],
                "description": "Opening balance",
                "dr": d("0"), "cr": abs(bal),
            })

    # --- JE-002 to JE-N: Bank Transactions ---
    for txn in all_txns:
        je_num += 1
        gl_acct = CLASSIFICATION_MAP.get(txn["acct_code"], txn["acct_code"])
        gl_info = COA.get(gl_acct, {"name": f"Unknown ({gl_acct})", "type": "Unknown"})
        amt = d(txn["amount"])
        desc = txn["description"][:80]

        if amt > 0:
            # Inflow: DR Bank, CR GL account
            journal_entries.append({
                "je_num": je_num, "date": txn["date"],
                "acct": "1000", "acct_name": COA["1000"]["name"],
                "description": desc, "dr": amt, "cr": d("0"),
            })
            journal_entries.append({
                "je_num": je_num, "date": txn["date"],
                "acct": gl_acct, "acct_name": gl_info["name"],
                "description": desc, "dr": d("0"), "cr": amt,
            })
        elif amt < 0:
            # Outflow: DR GL account, CR Bank
            abs_amt = abs(amt)
            journal_entries.append({
                "je_num": je_num, "date": txn["date"],
                "acct": gl_acct, "acct_name": gl_info["name"],
                "description": desc, "dr": abs_amt, "cr": d("0"),
            })
            journal_entries.append({
                "je_num": je_num, "date": txn["date"],
                "acct": "1000", "acct_name": COA["1000"]["name"],
                "description": desc, "dr": d("0"), "cr": abs_amt,
            })
        # Skip zero-amount transactions (no JE needed)

    # --- JE-N+1: Depreciation Music ---
    je_num += 1
    journal_entries.append({
        "je_num": je_num, "date": "2025-01-31",
        "acct": "5300", "acct_name": COA["5300"]["name"],
        "description": "Annual depreciation - Music System & Equipment @ 20% SL",
        "dr": d("12070.78"), "cr": d("0"),
    })
    journal_entries.append({
        "je_num": je_num, "date": "2025-01-31",
        "acct": "1399", "acct_name": COA["1399"]["name"],
        "description": "Annual depreciation - Music System & Equipment @ 20% SL",
        "dr": d("0"), "cr": d("12070.78"),
    })

    # --- JE-N+2: Depreciation Computer ---
    je_num += 1
    journal_entries.append({
        "je_num": je_num, "date": "2025-01-31",
        "acct": "5310", "acct_name": COA["5310"]["name"],
        "description": "Annual depreciation - Computer & POS @ 20% SL",
        "dr": d("713.60"), "cr": d("0"),
    })
    journal_entries.append({
        "je_num": je_num, "date": "2025-01-31",
        "acct": "1398", "acct_name": COA["1398"]["name"],
        "description": "Annual depreciation - Computer & POS @ 20% SL",
        "dr": d("0"), "cr": d("713.60"),
    })

    # --- JE-N+3: Prepayment Consumed ---
    je_num += 1
    journal_entries.append({
        "je_num": je_num, "date": "2024-05-31",
        "acct": "5970", "acct_name": COA["5970"]["name"],
        "description": "Prepayment consumed - Secretarial fee Feb-May 24",
        "dr": d("333.33"), "cr": d("0"),
    })
    journal_entries.append({
        "je_num": je_num, "date": "2024-05-31",
        "acct": "1210", "acct_name": COA["1210"]["name"],
        "description": "Prepayment consumed - Secretarial fee Feb-May 24",
        "dr": d("0"), "cr": d("333.33"),
    })

    # --- JE-N+4: Accrued Audit Fee ---
    je_num += 1
    journal_entries.append({
        "je_num": je_num, "date": "2025-01-31",
        "acct": "5520", "acct_name": COA["5520"]["name"],
        "description": "Accrued audit fee FY ending [FY_END]",
        "dr": d("3000.00"), "cr": d("0"),
    })
    journal_entries.append({
        "je_num": je_num, "date": "2025-01-31",
        "acct": "2100", "acct_name": COA["2100"]["name"],
        "description": "Accrued audit fee FY ending [FY_END]",
        "dr": d("0"), "cr": d("3000.00"),
    })

    # --- JE-N+5: Accrued Accounting Fee ---
    je_num += 1
    journal_entries.append({
        "je_num": je_num, "date": "2025-01-31",
        "acct": "5510", "acct_name": COA["5510"]["name"],
        "description": "Accrued accounting fee FY ending [FY_END]",
        "dr": d("2000.00"), "cr": d("0"),
    })
    journal_entries.append({
        "je_num": je_num, "date": "2025-01-31",
        "acct": "2100", "acct_name": COA["2100"]["name"],
        "description": "Accrued accounting fee FY ending [FY_END]",
        "dr": d("0"), "cr": d("2000.00"),
    })

    # --- Validate all JEs balance ---
    je_groups: dict[int, list[dict]] = defaultdict(list)
    for je in journal_entries:
        je_groups[je["je_num"]].append(je)

    for num, lines in je_groups.items():
        total_dr = sum(l["dr"] for l in lines)
        total_cr = sum(l["cr"] for l in lines)
        diff = abs(total_dr - total_cr)
        assert diff < Decimal("0.01"), f"JE-{num:04d} does not balance! DR={total_dr}, CR={total_cr}, diff={diff}"

    print(f"\nJournal entries: {je_num} JEs, {len(journal_entries)} lines")
    print("All JEs balance ✓")

    # ===================================================================
    # 2. GENERAL LEDGER COMPUTATION
    # ===================================================================

    gl: dict[str, list[dict]] = defaultdict(list)
    gl_balance: dict[str, Decimal] = defaultdict(lambda: d("0"))

    for je in journal_entries:
        acct = je["acct"]
        net = je["dr"] - je["cr"]
        gl_balance[acct] += net
        gl[acct].append({
            "je_num": je["je_num"],
            "date": je["date"],
            "description": je["description"],
            "dr": je["dr"],
            "cr": je["cr"],
            "balance": gl_balance[acct],
        })

    # ===================================================================
    # 3. TRIAL BALANCE
    # ===================================================================

    tb = OrderedDict()
    for acct in sorted(gl_balance.keys()):
        bal = gl_balance[acct]
        if abs(bal) < Decimal("0.01"):
            continue
        acct_info = COA.get(acct, {"name": f"Unknown ({acct})", "type": "Unknown"})
        dr_val = bal if bal > 0 else d("0")
        cr_val = abs(bal) if bal < 0 else d("0")
        tb[acct] = {
            "name": acct_info["name"],
            "type": acct_info["type"],
            "dr": dr_val,
            "cr": cr_val,
            "balance": bal,
        }

    tb_total_dr = sum(v["dr"] for v in tb.values())
    tb_total_cr = sum(v["cr"] for v in tb.values())
    tb_diff = abs(tb_total_dr - tb_total_cr)

    # ===================================================================
    # 4. INCOME STATEMENT COMPUTATION
    # ===================================================================

    revenue_accts = {k: v for k, v in tb.items() if k.startswith("4")}
    expense_accts = {k: v for k, v in tb.items() if k.startswith("5") or k.startswith("6")}

    total_revenue = sum(v["cr"] - v["dr"] for v in revenue_accts.values())
    cogs = tb.get("5050", {"dr": d("0"), "cr": d("0")})
    cogs_amt = cogs["dr"] - cogs["cr"]
    gross_profit = total_revenue - cogs_amt
    total_opex = sum(v["dr"] - v["cr"] for k, v in expense_accts.items() if k != "5050")
    net_profit = gross_profit - total_opex

    # ===================================================================
    # 5. BALANCE SHEET COMPUTATION
    # ===================================================================

    # Current Assets
    cash_bank = gl_balance.get("1000", d("0"))
    cash_hand = gl_balance.get("1010", d("0"))
    trade_recv = gl_balance.get("1100", d("0"))
    prepayments = gl_balance.get("1210", d("0"))
    deposits = gl_balance.get("1220", d("0"))
    total_ca = cash_bank + cash_hand + trade_recv + prepayments + deposits

    # Non-Current Assets (PPE net of accumulated depreciation)
    ppe_music_cost = gl_balance.get("1300", d("0"))
    ppe_comp_cost = gl_balance.get("1310", d("0"))
    accum_dep_music = gl_balance.get("1399", d("0"))   # negative
    accum_dep_comp = gl_balance.get("1398", d("0"))     # negative
    ppe_music_net = ppe_music_cost + accum_dep_music
    ppe_comp_net = ppe_comp_cost + accum_dep_comp
    total_nca = ppe_music_net + ppe_comp_net
    total_assets = total_ca + total_nca

    # Current Liabilities
    trade_pay = abs(gl_balance.get("2000", d("0")))
    other_pay = abs(gl_balance.get("2100", d("0")))
    epf_pay = abs(gl_balance.get("2110", d("0")))
    socso_pay = abs(gl_balance.get("2120", d("0")))
    eis_pay = abs(gl_balance.get("2130", d("0")))
    due_director = abs(gl_balance.get("2350", d("0")))
    due_shareholders = abs(gl_balance.get("2360", d("0")))
    suspense = abs(gl_balance.get("2900", d("0")))
    total_cl = trade_pay + other_pay + epf_pay + socso_pay + eis_pay + due_director + due_shareholders + suspense

    # Equity
    share_cap = abs(gl_balance.get("3000", d("0")))
    retained = gl_balance.get("3200", d("0"))  # DR balance (loss) = positive
    total_equity = share_cap - retained + net_profit  # 1000 - 58547.35 + net_profit

    # BS check
    total_eq_liab = total_equity + total_cl
    bs_diff = abs(total_assets - total_eq_liab)

    # ===================================================================
    # 6. CONSOLE OUTPUT
    # ===================================================================

    print("\n" + "=" * 70)
    print("[CLIENT_NAME] — TRIAL BALANCE as at 31 January 2025")
    print("=" * 70)
    print(f"{'Acct':<6} {'Account Name':<45} {'DR':>14} {'CR':>14}")
    print("-" * 80)
    for acct, info in tb.items():
        dr_str = f"{info['dr']:,.2f}" if info['dr'] != d("0") else ""
        cr_str = f"{info['cr']:,.2f}" if info['cr'] != d("0") else ""
        print(f"{acct:<6} {info['name']:<45} {dr_str:>14} {cr_str:>14}")
    print("-" * 80)
    print(f"{'TOTAL':<52} {tb_total_dr:>14,.2f} {tb_total_cr:>14,.2f}")
    print(f"{'DIFFERENCE':<52} {tb_diff:>14,.2f}")
    if tb_diff < Decimal("0.01"):
        print("TB Difference: 0.00 ✓ BALANCED")
    else:
        print(f"TB Difference: {tb_diff:,.2f} ✗ UNBALANCED")

    # --- Income Statement ---
    print("\n" + "=" * 70)
    print("INCOME STATEMENT for FY ended 31 January 2025")
    print("=" * 70)

    print("\nREVENUE")
    for acct, info in revenue_accts.items():
        amt = info["cr"] - info["dr"]
        print(f"  {acct} {info['name']:<45} {amt:>14,.2f}")
    print(f"  {'Total Revenue':<51} {total_revenue:>14,.2f}")

    print(f"\nCOST OF GOODS SOLD")
    print(f"  5050 {'Food & Beverage Cost':<45} ({cogs_amt:>12,.2f})")
    print(f"  {'GROSS PROFIT':<51} {gross_profit:>14,.2f}")

    print(f"\nOPERATING EXPENSES")
    for acct, info in expense_accts.items():
        if acct == "5050":
            continue
        amt = info["dr"] - info["cr"]
        print(f"  {acct} {info['name']:<45} ({amt:>12,.2f})")
    print(f"  {'Total Operating Expenses':<51} ({total_opex:>12,.2f})")

    print(f"\n  {'NET PROFIT / (LOSS)':<51} {net_profit:>14,.2f}")

    # --- Balance Sheet ---
    print("\n" + "=" * 70)
    print("BALANCE SHEET as at 31 January 2025")
    print("=" * 70)

    print("\nCURRENT ASSETS")
    print(f"  Cash at Bank                                       {cash_bank:>14,.2f}")
    print(f"  Cash in Hand                                       {cash_hand:>14,.2f}")
    print(f"  Trade Receivables                                  {trade_recv:>14,.2f}")
    print(f"  Prepayments                                        {prepayments:>14,.2f}")
    print(f"  Deposits                                           {deposits:>14,.2f}")
    print(f"  {'Total Current Assets':<51} {total_ca:>14,.2f}")

    print(f"\nNON-CURRENT ASSETS")
    print(f"  PPE - Music (Cost {ppe_music_cost:,.2f}, Dep {accum_dep_music:,.2f})  Net: {ppe_music_net:>10,.2f}")
    print(f"  PPE - Computer (Cost {ppe_comp_cost:,.2f}, Dep {accum_dep_comp:,.2f})  Net: {ppe_comp_net:>10,.2f}")
    print(f"  {'Total Non-Current Assets':<51} {total_nca:>14,.2f}")

    print(f"\n  {'TOTAL ASSETS':<51} {total_assets:>14,.2f}")

    print(f"\nCURRENT LIABILITIES")
    print(f"  Trade Payables                                     {trade_pay:>14,.2f}")
    print(f"  Other Payables & Accruals                          {other_pay:>14,.2f}")
    if epf_pay > 0:
        print(f"  EPF Payable                                        {epf_pay:>14,.2f}")
    if socso_pay > 0:
        print(f"  SOCSO Payable                                      {socso_pay:>14,.2f}")
    if eis_pay > 0:
        print(f"  EIS Payable                                        {eis_pay:>14,.2f}")
    print(f"  Amount Due to Director                             {due_director:>14,.2f}")
    print(f"  Amount Due to Shareholders                         {due_shareholders:>14,.2f}")
    if suspense > 0:
        print(f"  Suspense Account                                   {suspense:>14,.2f}")
    print(f"  {'Total Current Liabilities':<51} {total_cl:>14,.2f}")

    print(f"\nEQUITY")
    print(f"  Share Capital                                      {share_cap:>14,.2f}")
    print(f"  Retained Earnings (Prior Years)                    ({retained:>12,.2f})")
    print(f"  Net Profit / (Loss) for the Year                   {net_profit:>14,.2f}")
    print(f"  {'Total Equity':<51} {total_equity:>14,.2f}")

    print(f"\n  {'TOTAL EQUITY + LIABILITIES':<51} {total_eq_liab:>14,.2f}")
    print(f"\n  BS CHECK: Assets {total_assets:,.2f} vs Eq+Liab {total_eq_liab:,.2f}")
    print(f"  BS Difference: {bs_diff:,.2f}", end="")
    if bs_diff < Decimal("0.01"):
        print(" ✓")
    else:
        print(" ✗ UNBALANCED")

    # --- Bank Closing Balance Check ---
    print("\n" + "=" * 70)
    print("BANK RECONCILIATION CHECK")
    print("=" * 70)
    bank_closing = gl_balance.get("1000", d("0"))
    bank_stmt = d("12511.57")
    bank_diff = bank_closing - bank_stmt
    print(f"Bank GL closing: {bank_closing:,.2f}")
    print(f"Bank stmt closing: {bank_stmt:,.2f}")
    print(f"Bank difference: {bank_diff:,.2f}")

    # ===================================================================
    # EXCEL WRITING
    # ===================================================================

    print("\n" + "=" * 70)
    print("Writing Excel workbook...")

    wb = Workbook()
    tracker = RowTracker()

    HEADER_FONT = Font(bold=True, size=11, name="Arial")
    TITLE_FONT = Font(bold=True, size=14, name="Arial")
    SUBTITLE_FONT = Font(bold=True, size=12, name="Arial")
    NORMAL_FONT = Font(size=10, name="Arial")
    CURRENCY_FMT = '#,##0.00;(#,##0.00)'
    CHECK_OK_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    CHECK_FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    THIN_BORDER = Border(bottom=Side(style='thin'))
    DOUBLE_BORDER = Border(top=Side(style='double'), bottom=Side(style='double'))

    # --- Pre-create all sheet tabs ---
    ws_info = wb.active
    ws_info.title = "Company Info"
    ws_coa = wb.create_sheet("Chart of Accounts")
    ws_je = wb.create_sheet("Journal Entries")
    ws_gl = wb.create_sheet("General Ledger")
    ws_bank = wb.create_sheet("Bank Transactions")
    ws_tb = wb.create_sheet("Trial Balance")
    ws_is = wb.create_sheet("Income Statement")
    ws_bs = wb.create_sheet("Balance Sheet")
    ws_far = wb.create_sheet("Fixed Assets")
    ws_payroll = wb.create_sheet("Payroll Summary")
    ws_tax = wb.create_sheet("Tax Computation")
    ws_queries = wb.create_sheet("Queries & Notes")

    # ===================================================================
    # SHEET 1: COMPANY INFO
    # ===================================================================
    print("  Sheet 1: Company Info...")

    ws_info.column_dimensions["A"].width = 25
    ws_info.column_dimensions["B"].width = 80

    info_data = [
        ("[CLIENT_NAME]", ""),
        ("Registration No.", "[SSM_REGISTRATION]"),
        ("", ""),
        ("Financial Year:", "1 February 2024 to 31 January 2025"),
        ("Entity Type:", "Sdn Bhd (Private Limited)"),
        ("Framework:", "MPERS"),
        ("MSIC Code:", "56210 - Event/food caterers"),
        ("Tax File No:", "C60033841100"),
        ("Principal Activity:", "Food & beverage, event management, catering"),
        ("Registered Office:", "Suite 9B.05, Level 9B, Wisma E&C, No 2 Lorong Dungun Kiri, Damansara Heights, 50490 KL"),
        ("Business Address:", "G06, Ground Floor, same building"),
        ("Bank Account:", "[BANK_NAME] [ACCOUNT_NUMBER]"),
        ("Auditor (Prior Year):", "LKT & Associates (AF 002169)"),
        ("", ""),
        ("Directors:", ""),
        ("", "Mohammad Luqman Nur Hakim Bin Md Zim (Service Director)"),
        ("", "Wan Saifulruddin Bin Wan Jan (Service Director, 35% shares)"),
        ("", ""),
        ("Shareholders:", ""),
        ("", "Ahmad Helmi Bin Azhar - 60%"),
        ("", "Wan Saifulruddin Bin Wan Jan - 35%"),
        ("", "Gurdial Singh Nijar - 5%"),
        ("", ""),
        ("Going Concern:", "See Queries & Notes sheet for going concern assessment"),
        ("", ""),
        ("Prepared by:", "[FIRM_NAME] Chartered Accountants ([CA_REGISTRATION])"),
        ("Date:", date.today().strftime("%d %B %Y")),
    ]

    for r, (a, b) in enumerate(info_data, start=1):
        ws_info.cell(row=r, column=1, value=a).font = HEADER_FONT if a else NORMAL_FONT
        ws_info.cell(row=r, column=2, value=b).font = NORMAL_FONT

    ws_info.cell(row=1, column=1).font = TITLE_FONT

    # ===================================================================
    # SHEET 2: CHART OF ACCOUNTS
    # ===================================================================
    print("  Sheet 2: Chart of Accounts...")

    ws_coa.column_dimensions["A"].width = 8
    ws_coa.column_dimensions["B"].width = 55
    ws_coa.column_dimensions["C"].width = 12
    ws_coa.column_dimensions["D"].width = 8

    coa_headers = ["Code", "Account Name", "Type", "Normal Balance"]
    for c, h in enumerate(coa_headers, start=1):
        cell = ws_coa.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT

    for r, (code, info) in enumerate(COA.items(), start=2):
        ws_coa.cell(row=r, column=1, value=code).font = NORMAL_FONT
        ws_coa.cell(row=r, column=2, value=info["name"]).font = NORMAL_FONT
        ws_coa.cell(row=r, column=3, value=info["type"]).font = NORMAL_FONT
        ws_coa.cell(row=r, column=4, value=info["normal"]).font = NORMAL_FONT

    # ===================================================================
    # SHEET 3: JOURNAL ENTRIES — WITH CHECK FORMULAS
    # ===================================================================
    print("  Sheet 3: Journal Entries...")

    ws_je.column_dimensions["A"].width = 6
    ws_je.column_dimensions["B"].width = 12
    ws_je.column_dimensions["C"].width = 8
    ws_je.column_dimensions["D"].width = 40
    ws_je.column_dimensions["E"].width = 50
    ws_je.column_dimensions["F"].width = 14
    ws_je.column_dimensions["G"].width = 14
    ws_je.column_dimensions["H"].width = 14

    # Title rows
    ws_je.cell(row=1, column=1, value="[CLIENT_NAME]").font = TITLE_FONT
    ws_je.cell(row=2, column=1, value="Journal Entries — FY ended 31 January 2025").font = SUBTITLE_FONT

    # Headers on row 4
    je_headers = ["JE#", "Date", "Acct Code", "Account Name", "Description", "Debit", "Credit", "CHECK"]
    for c, h in enumerate(je_headers, start=1):
        cell = ws_je.cell(row=4, column=c, value=h)
        cell.font = HEADER_FONT

    # Write JE lines starting at row 5
    je_row = 5
    je_group_start: dict[int, int] = {}  # je_num -> first row
    je_group_end: dict[int, int] = {}    # je_num -> last row

    for je in journal_entries:
        num = je["je_num"]
        if num not in je_group_start:
            je_group_start[num] = je_row

        ws_je.cell(row=je_row, column=1, value=num).font = NORMAL_FONT
        ws_je.cell(row=je_row, column=2, value=je["date"]).font = NORMAL_FONT
        ws_je.cell(row=je_row, column=3, value=je["acct"]).font = NORMAL_FONT
        ws_je.cell(row=je_row, column=4, value=je["acct_name"]).font = NORMAL_FONT
        ws_je.cell(row=je_row, column=5, value=je["description"]).font = NORMAL_FONT

        dr_cell = ws_je.cell(row=je_row, column=6, value=float(je["dr"]))
        dr_cell.number_format = CURRENCY_FMT
        dr_cell.font = NORMAL_FONT

        cr_cell = ws_je.cell(row=je_row, column=7, value=float(je["cr"]))
        cr_cell.number_format = CURRENCY_FMT
        cr_cell.font = NORMAL_FONT

        je_group_end[num] = je_row
        je_row += 1

    # Write CHECK formulas on the last row of each JE group
    for num, last_row in je_group_end.items():
        first_row = je_group_start[num]
        check_formula = f"=SUM(F{first_row}:F{last_row})-SUM(G{first_row}:G{last_row})"
        check_cell = ws_je.cell(row=last_row, column=8, value=check_formula)
        check_cell.number_format = CURRENCY_FMT
        check_cell.font = NORMAL_FONT

    # Grand totals
    je_last_data_row = je_row - 1
    total_row = je_row + 1
    ws_je.cell(row=total_row, column=5, value="TOTAL").font = HEADER_FONT
    total_dr_cell = ws_je.cell(row=total_row, column=6, value=f"=SUM(F5:F{je_last_data_row})")
    total_dr_cell.number_format = CURRENCY_FMT
    total_dr_cell.font = HEADER_FONT
    total_dr_cell.border = DOUBLE_BORDER

    total_cr_cell = ws_je.cell(row=total_row, column=7, value=f"=SUM(G5:G{je_last_data_row})")
    total_cr_cell.number_format = CURRENCY_FMT
    total_cr_cell.font = HEADER_FONT
    total_cr_cell.border = DOUBLE_BORDER

    total_check_cell = ws_je.cell(row=total_row, column=8, value=f"=F{total_row}-G{total_row}")
    total_check_cell.number_format = CURRENCY_FMT
    total_check_cell.font = HEADER_FONT

    print(f"    {len(journal_entries)} JE lines written, {len(je_group_end)} JE groups with CHECK formulas")

    # ===================================================================
    # SHEET 4: GENERAL LEDGER — WITH RUNNING BALANCE FORMULAS
    # ===================================================================
    print("  Sheet 4: General Ledger...")

    ws_gl.column_dimensions["A"].width = 12
    ws_gl.column_dimensions["B"].width = 8
    ws_gl.column_dimensions["C"].width = 50
    ws_gl.column_dimensions["D"].width = 14
    ws_gl.column_dimensions["E"].width = 14
    ws_gl.column_dimensions["F"].width = 14

    gl_row = 1
    gl_closing_rows: dict[str, int] = {}  # acct_code -> closing balance row

    for acct_code in sorted(gl.keys()):
        entries = gl[acct_code]
        acct_info = COA.get(acct_code, {"name": f"Unknown ({acct_code})"})
        acct_name = acct_info["name"]

        # Account header row
        header_cell = ws_gl.cell(row=gl_row, column=1, value=f"{acct_code} — {acct_name}")
        header_cell.font = HEADER_FONT
        ws_gl.merge_cells(start_row=gl_row, start_column=1, end_row=gl_row, end_column=3)
        gl_row += 1

        # Column sub-headers
        sub_headers = ["Date", "JE#", "Description", "Debit", "Credit", "Balance"]
        for c, h in enumerate(sub_headers, start=1):
            cell = ws_gl.cell(row=gl_row, column=c, value=h)
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
        gl_row += 1

        # GL entries
        first_entry_row = gl_row
        for i, entry in enumerate(entries):
            ws_gl.cell(row=gl_row, column=1, value=entry["date"]).font = NORMAL_FONT
            ws_gl.cell(row=gl_row, column=2, value=entry["je_num"]).font = NORMAL_FONT
            ws_gl.cell(row=gl_row, column=3, value=entry["description"]).font = NORMAL_FONT

            # Debit
            dr_val = float(entry["dr"]) if entry["dr"] > 0 else None
            dr_cell = ws_gl.cell(row=gl_row, column=4, value=dr_val)
            dr_cell.number_format = CURRENCY_FMT
            dr_cell.font = NORMAL_FONT

            # Credit
            cr_val = float(entry["cr"]) if entry["cr"] > 0 else None
            cr_cell = ws_gl.cell(row=gl_row, column=5, value=cr_val)
            cr_cell.number_format = CURRENCY_FMT
            cr_cell.font = NORMAL_FONT

            # Balance formula
            if i == 0:
                bal_formula = f"=IF(ISNUMBER(D{gl_row}),D{gl_row},0)-IF(ISNUMBER(E{gl_row}),E{gl_row},0)"
            else:
                prev_row = gl_row - 1
                bal_formula = f"=F{prev_row}+IF(ISNUMBER(D{gl_row}),D{gl_row},0)-IF(ISNUMBER(E{gl_row}),E{gl_row},0)"
            bal_cell = ws_gl.cell(row=gl_row, column=6, value=bal_formula)
            bal_cell.number_format = CURRENCY_FMT
            bal_cell.font = NORMAL_FONT

            gl_row += 1

        # Closing balance summary row
        last_entry_row = gl_row - 1
        ws_gl.cell(row=gl_row, column=3, value="Closing Balance").font = HEADER_FONT
        closing_cell = ws_gl.cell(row=gl_row, column=6, value=f"=F{last_entry_row}")
        closing_cell.number_format = CURRENCY_FMT
        closing_cell.font = HEADER_FONT
        closing_cell.border = DOUBLE_BORDER

        gl_closing_rows[acct_code] = gl_row
        tracker.record("General Ledger", acct_code, gl_row)
        gl_row += 2  # blank row between accounts

    # --- GL SUMMARY BLOCK ---
    gl_row += 1
    ws_gl.cell(row=gl_row, column=1, value="GENERAL LEDGER SUMMARY").font = SUBTITLE_FONT
    gl_row += 1

    summary_headers = ["Code", "Account Name", "Closing Balance"]
    for c, h in enumerate(summary_headers, start=1):
        cell = ws_gl.cell(row=gl_row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    gl_row += 1

    for acct_code in sorted(gl.keys()):
        acct_info = COA.get(acct_code, {"name": f"Unknown ({acct_code})"})
        ws_gl.cell(row=gl_row, column=1, value=acct_code).font = NORMAL_FONT
        ws_gl.cell(row=gl_row, column=2, value=acct_info["name"]).font = NORMAL_FONT

        closing_row = gl_closing_rows[acct_code]
        summary_cell = ws_gl.cell(row=gl_row, column=3, value=f"=F{closing_row}")
        summary_cell.number_format = CURRENCY_FMT
        summary_cell.font = NORMAL_FONT

        tracker.record("GL Summary", acct_code, gl_row)
        gl_row += 1

    print(f"    {len(gl)} accounts written with running balance formulas")
    print(f"    GL Summary block at rows {gl_row - len(gl)} to {gl_row - 1}")

    # ===================================================================
    # SHEET 5: BANK TRANSACTIONS — WITH RUNNING BALANCE FORMULAS
    # ===================================================================
    print("  Sheet 5: Bank Transactions...")

    ws_bank.column_dimensions["A"].width = 12
    ws_bank.column_dimensions["B"].width = 60
    ws_bank.column_dimensions["C"].width = 14
    ws_bank.column_dimensions["D"].width = 14
    ws_bank.column_dimensions["E"].width = 14
    ws_bank.column_dimensions["F"].width = 10
    ws_bank.column_dimensions["G"].width = 35

    # Row 1: Headers
    bank_headers = ["Date", "Description", "Credit (In)", "Debit (Out)", "Balance", "Acct Code", "Acct Name"]
    for c, h in enumerate(bank_headers, start=1):
        cell = ws_bank.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT

    # Row 2: Opening balance
    ws_bank.cell(row=2, column=1, value="2024-02-01").font = NORMAL_FONT
    ws_bank.cell(row=2, column=2, value="OPENING BALANCE").font = HEADER_FONT
    ws_bank.cell(row=2, column=3, value=0).number_format = CURRENCY_FMT
    ws_bank.cell(row=2, column=4, value=0).number_format = CURRENCY_FMT
    ob_cell = ws_bank.cell(row=2, column=5, value=15266.41)
    ob_cell.number_format = CURRENCY_FMT
    ob_cell.font = HEADER_FONT

    # Row 3 onwards: transactions
    bank_row = 3
    for txn in all_txns:
        amt = d(txn["amount"])
        gl_acct = CLASSIFICATION_MAP.get(txn["acct_code"], txn["acct_code"])
        gl_info = COA.get(gl_acct, {"name": f"Unknown ({gl_acct})"})

        ws_bank.cell(row=bank_row, column=1, value=txn["date"]).font = NORMAL_FONT
        ws_bank.cell(row=bank_row, column=2, value=txn["description"][:80]).font = NORMAL_FONT

        # Credit (In) — positive amounts
        credit_val = float(amt) if amt > 0 else 0
        cr_cell = ws_bank.cell(row=bank_row, column=3, value=credit_val)
        cr_cell.number_format = CURRENCY_FMT
        cr_cell.font = NORMAL_FONT

        # Debit (Out) — negative amounts as positive
        debit_val = float(abs(amt)) if amt < 0 else 0
        dr_cell = ws_bank.cell(row=bank_row, column=4, value=debit_val)
        dr_cell.number_format = CURRENCY_FMT
        dr_cell.font = NORMAL_FONT

        # Balance formula
        prev_row = bank_row - 1
        bal_formula = f"=E{prev_row}+C{bank_row}-D{bank_row}"
        bal_cell = ws_bank.cell(row=bank_row, column=5, value=bal_formula)
        bal_cell.number_format = CURRENCY_FMT
        bal_cell.font = NORMAL_FONT

        ws_bank.cell(row=bank_row, column=6, value=gl_acct).font = NORMAL_FONT
        ws_bank.cell(row=bank_row, column=7, value=gl_info["name"]).font = NORMAL_FONT

        bank_row += 1

    last_txn_row = bank_row - 1
    tracker.record("Bank Transactions", "closing", last_txn_row)

    # Bank reconciliation block
    bank_row += 1
    ws_bank.cell(row=bank_row, column=2, value="Bank Statement Closing Balance ([FY_END]):").font = HEADER_FONT
    bank_stmt_cell = ws_bank.cell(row=bank_row, column=5, value=12511.57)
    bank_stmt_cell.number_format = CURRENCY_FMT
    bank_stmt_cell.font = HEADER_FONT
    bank_stmt_row = bank_row

    bank_row += 1
    ws_bank.cell(row=bank_row, column=2, value="GL Closing Balance:").font = HEADER_FONT
    gl_closing_cell = ws_bank.cell(row=bank_row, column=5, value=f"=E{last_txn_row}")
    gl_closing_cell.number_format = CURRENCY_FMT
    gl_closing_cell.font = HEADER_FONT

    bank_row += 1
    ws_bank.cell(row=bank_row, column=2, value="Difference:").font = HEADER_FONT
    diff_cell = ws_bank.cell(row=bank_row, column=5, value=f"=E{last_txn_row}-E{bank_stmt_row}")
    diff_cell.number_format = CURRENCY_FMT
    diff_cell.font = HEADER_FONT

    print(f"    {len(all_txns)} bank transactions written")
    print(f"    Bank reconciliation block at row {bank_stmt_row}")

    # ===================================================================
    # SHEET 6: TRIAL BALANCE — references GL Summary
    # ===================================================================
    print("  Sheet 6: Trial Balance...")

    ws_tb.column_dimensions["A"].width = 8
    ws_tb.column_dimensions["B"].width = 55
    ws_tb.column_dimensions["C"].width = 16
    ws_tb.column_dimensions["D"].width = 16

    ws_tb.cell(row=1, column=1, value="[CLIENT_NAME]").font = TITLE_FONT
    ws_tb.cell(row=2, column=1, value="TRIAL BALANCE AS AT 31 JANUARY 2025").font = SUBTITLE_FONT

    tb_headers = ["Code", "Account", "Debit (RM)", "Credit (RM)"]
    for c, h in enumerate(tb_headers, start=1):
        cell = ws_tb.cell(row=4, column=c, value=h)
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    tb_row = 5
    for acct_code, info in tb.items():
        ws_tb.cell(row=tb_row, column=1, value=acct_code).font = NORMAL_FONT
        ws_tb.cell(row=tb_row, column=2, value=info["name"]).font = NORMAL_FONT

        # GL Summary closing balance is in column C of the General Ledger sheet
        gl_summary_row = tracker.row("GL Summary", acct_code)

        # DR column: if GL Summary closing balance > 0, show it; else 0
        dr_formula = f"=IF('General Ledger'!C{gl_summary_row}>0,'General Ledger'!C{gl_summary_row},0)"
        dr_cell = ws_tb.cell(row=tb_row, column=3, value=dr_formula)
        dr_cell.number_format = CURRENCY_FMT
        dr_cell.font = NORMAL_FONT

        # CR column: if GL Summary closing balance < 0, show absolute value; else 0
        cr_formula = f"=IF('General Ledger'!C{gl_summary_row}<0,ABS('General Ledger'!C{gl_summary_row}),0)"
        cr_cell = ws_tb.cell(row=tb_row, column=4, value=cr_formula)
        cr_cell.number_format = CURRENCY_FMT
        cr_cell.font = NORMAL_FONT

        tracker.record("Trial Balance", acct_code, tb_row)
        tb_row += 1

    tb_last_data_row = tb_row - 1

    # Blank row
    tb_row += 1

    # Total row
    ws_tb.cell(row=tb_row, column=2, value="TOTAL").font = HEADER_FONT
    total_dr_tb = ws_tb.cell(row=tb_row, column=3, value=f"=SUM(C5:C{tb_last_data_row})")
    total_dr_tb.number_format = CURRENCY_FMT
    total_dr_tb.font = HEADER_FONT
    total_dr_tb.border = DOUBLE_BORDER

    total_cr_tb = ws_tb.cell(row=tb_row, column=4, value=f"=SUM(D5:D{tb_last_data_row})")
    total_cr_tb.number_format = CURRENCY_FMT
    total_cr_tb.font = HEADER_FONT
    total_cr_tb.border = DOUBLE_BORDER

    tb_total_row = tb_row
    tb_row += 1

    # CHECK row
    ws_tb.cell(row=tb_row, column=2, value="CHECK (DR - CR)").font = HEADER_FONT
    tb_check_cell = ws_tb.cell(row=tb_row, column=3, value=f"=C{tb_total_row}-D{tb_total_row}")
    tb_check_cell.number_format = CURRENCY_FMT
    tb_check_cell.font = HEADER_FONT
    tracker.record("Trial Balance", "check", tb_row)

    print(f"    {len(tb)} accounts written with GL Summary cross-references")

    # ===================================================================
    # SHEET 7: INCOME STATEMENT — references Trial Balance
    # ===================================================================
    print("  Sheet 7: Income Statement...")

    ws_is.column_dimensions["A"].width = 45
    ws_is.column_dimensions["B"].width = 40
    ws_is.column_dimensions["C"].width = 16

    ws_is.cell(row=1, column=1, value="[CLIENT_NAME]").font = TITLE_FONT
    ws_is.cell(row=2, column=1, value="STATEMENT OF COMPREHENSIVE INCOME").font = SUBTITLE_FONT
    ws_is.cell(row=3, column=1, value="FOR THE YEAR ENDED 31 JANUARY 2025").font = SUBTITLE_FONT

    ws_is.cell(row=5, column=3, value="RM").font = HEADER_FONT

    is_row = 6

    # --- REVENUE ---
    ws_is.cell(row=is_row, column=1, value="REVENUE").font = HEADER_FONT
    is_row += 1

    revenue_codes = [
        ("4000", "Event Revenue"),
        ("4100", "F&B / Cafe Revenue"),
        ("4200", "Other Income"),
    ]
    is_revenue_first = is_row
    for acct_code, label in revenue_codes:
        if tracker.has("Trial Balance", acct_code):
            tb_r = tracker.row("Trial Balance", acct_code)
            # Revenue = CR column (D) of Trial Balance
            rev_formula = f"='Trial Balance'!D{tb_r}"
            ws_is.cell(row=is_row, column=2, value=label).font = NORMAL_FONT
            rev_cell = ws_is.cell(row=is_row, column=3, value=rev_formula)
            rev_cell.number_format = CURRENCY_FMT
            rev_cell.font = NORMAL_FONT
            tracker.record("Income Statement", acct_code, is_row)
            is_row += 1
    is_revenue_last = is_row - 1

    # Total Revenue
    ws_is.cell(row=is_row, column=2, value="TOTAL REVENUE").font = HEADER_FONT
    total_rev_cell = ws_is.cell(row=is_row, column=3, value=f"=SUM(C{is_revenue_first}:C{is_revenue_last})")
    total_rev_cell.number_format = CURRENCY_FMT
    total_rev_cell.font = HEADER_FONT
    total_rev_cell.border = THIN_BORDER
    total_rev_row = is_row
    tracker.record("Income Statement", "total_revenue", is_row)
    is_row += 2  # blank row

    # --- COST OF SALES ---
    ws_is.cell(row=is_row, column=1, value="COST OF SALES").font = HEADER_FONT
    is_row += 1

    cogs_row = None
    if tracker.has("Trial Balance", "5050"):
        tb_r = tracker.row("Trial Balance", "5050")
        # COGS = negate DR column (show as negative)
        cogs_formula = f"=-'Trial Balance'!C{tb_r}"
        ws_is.cell(row=is_row, column=2, value="Food & Beverage Cost").font = NORMAL_FONT
        cogs_cell = ws_is.cell(row=is_row, column=3, value=cogs_formula)
        cogs_cell.number_format = CURRENCY_FMT
        cogs_cell.font = NORMAL_FONT
        cogs_row = is_row
        tracker.record("Income Statement", "5050", is_row)
        is_row += 1
    is_row += 1  # blank row

    # GROSS PROFIT
    ws_is.cell(row=is_row, column=1, value="GROSS PROFIT").font = HEADER_FONT
    if cogs_row:
        gp_formula = f"=C{total_rev_row}+C{cogs_row}"
    else:
        gp_formula = f"=C{total_rev_row}"
    gp_cell = ws_is.cell(row=is_row, column=3, value=gp_formula)
    gp_cell.number_format = CURRENCY_FMT
    gp_cell.font = HEADER_FONT
    gp_cell.border = THIN_BORDER
    gross_profit_row = is_row
    tracker.record("Income Statement", "gross_profit", is_row)
    is_row += 2  # blank row

    # --- ADMINISTRATIVE & OPERATING EXPENSES ---
    ws_is.cell(row=is_row, column=1, value="ADMINISTRATIVE & OPERATING EXPENSES").font = HEADER_FONT
    is_row += 1

    expense_codes = [
        ("5000", "Salary & Wages"),
        ("5030", "Performance & Show Expenses"),
        ("5200", "Rental Expense"),
        ("5210", "Utilities"),
        ("5300", "Depreciation - Music Equipment"),
        ("5310", "Depreciation - Computer & POS"),
        ("5400", "Software & Subscriptions"),
        ("5510", "Accounting Fee (Accrued)"),
        ("5520", "Audit Fee (Accrued)"),
        ("5950", "Printing & Stationery"),
        ("5960", "Repairs & Maintenance"),
        ("5970", "Miscellaneous Expenses"),
        ("5980", "Bank Charges"),
    ]

    expense_cell_refs = []
    is_expense_first = is_row
    for acct_code, label in expense_codes:
        if tracker.has("Trial Balance", acct_code):
            tb_r = tracker.row("Trial Balance", acct_code)
            # Expense = negate DR column (show as negative)
            exp_formula = f"=-'Trial Balance'!C{tb_r}"
            ws_is.cell(row=is_row, column=2, value=label).font = NORMAL_FONT
            exp_cell = ws_is.cell(row=is_row, column=3, value=exp_formula)
            exp_cell.number_format = CURRENCY_FMT
            exp_cell.font = NORMAL_FONT
            expense_cell_refs.append(f"C{is_row}")
            tracker.record("Income Statement", acct_code, is_row)

            # Record specific accounts needed for tax computation
            if acct_code == "5300":
                tracker.record("Income Statement", "depr_music", is_row)
            elif acct_code == "5310":
                tracker.record("Income Statement", "depr_comp", is_row)
            elif acct_code == "5520":
                tracker.record("Income Statement", "audit_fee", is_row)
            elif acct_code == "5510":
                tracker.record("Income Statement", "accounting_fee", is_row)

            is_row += 1
    is_expense_last = is_row - 1

    # Total Expenses
    is_row += 1  # blank row before total
    ws_is.cell(row=is_row, column=2, value="TOTAL EXPENSES").font = HEADER_FONT
    total_exp_cell = ws_is.cell(row=is_row, column=3, value=f"=SUM(C{is_expense_first}:C{is_expense_last})")
    total_exp_cell.number_format = CURRENCY_FMT
    total_exp_cell.font = HEADER_FONT
    total_exp_cell.border = THIN_BORDER
    total_expenses_row = is_row
    tracker.record("Income Statement", "total_expenses", is_row)
    is_row += 2  # blank row

    # NET PROFIT / (LOSS) BEFORE TAX
    ws_is.cell(row=is_row, column=1, value="NET PROFIT / (LOSS) BEFORE TAX").font = Font(bold=True, size=12, name="Arial")
    np_formula = f"=C{gross_profit_row}+C{total_expenses_row}"
    np_cell = ws_is.cell(row=is_row, column=3, value=np_formula)
    np_cell.number_format = CURRENCY_FMT
    np_cell.font = Font(bold=True, size=12, name="Arial")
    np_cell.border = DOUBLE_BORDER
    tracker.record("Income Statement", "net_profit", is_row)

    print(f"    Revenue: {len([c for c, _ in revenue_codes if tracker.has('Trial Balance', c)])} accounts")
    print(f"    Expenses: {len(expense_cell_refs)} accounts")
    print(f"    Net profit formula at row {is_row}")

    # ===================================================================
    # SHEET 8: BALANCE SHEET — references TB, IS, and Fixed Assets
    # ===================================================================
    print("  Sheet 8: Balance Sheet...")

    ws_bs.column_dimensions["A"].width = 45
    ws_bs.column_dimensions["B"].width = 18

    ws_bs.cell(row=1, column=1, value="[CLIENT_NAME]").font = TITLE_FONT
    ws_bs.cell(row=2, column=1, value="STATEMENT OF FINANCIAL POSITION").font = SUBTITLE_FONT
    ws_bs.cell(row=3, column=1, value="AS AT 31 JANUARY 2025").font = SUBTITLE_FONT

    ws_bs.cell(row=5, column=2, value="RM").font = HEADER_FONT

    bs_row = 7

    # --- NON-CURRENT ASSETS ---
    ws_bs.cell(row=bs_row, column=1, value="NON-CURRENT ASSETS").font = HEADER_FONT
    bs_row += 1

    # Music System & Equipment (net) — write computed value; Task 5 will overwrite with FAR formula
    ws_bs.cell(row=bs_row, column=1, value="Music System & Equipment (net)").font = NORMAL_FONT
    nca_music_cell = ws_bs.cell(row=bs_row, column=2, value=float(ppe_music_net))
    nca_music_cell.number_format = CURRENCY_FMT
    nca_music_cell.font = NORMAL_FONT
    nca_music_row = bs_row
    tracker.record("Balance Sheet", "nca_music", bs_row)
    bs_row += 1

    # Computer & POS (net) — write computed value; Task 5 will overwrite with FAR formula
    ws_bs.cell(row=bs_row, column=1, value="Computer & POS (net)").font = NORMAL_FONT
    nca_comp_cell = ws_bs.cell(row=bs_row, column=2, value=float(ppe_comp_net))
    nca_comp_cell.number_format = CURRENCY_FMT
    nca_comp_cell.font = NORMAL_FONT
    nca_comp_row = bs_row
    tracker.record("Balance Sheet", "nca_comp", bs_row)
    bs_row += 1

    # Total Non-Current Assets
    ws_bs.cell(row=bs_row, column=1, value="Total Non-Current Assets").font = HEADER_FONT
    total_nca_cell = ws_bs.cell(row=bs_row, column=2, value=f"=SUM(B{nca_music_row}:B{nca_comp_row})")
    total_nca_cell.number_format = CURRENCY_FMT
    total_nca_cell.font = HEADER_FONT
    total_nca_cell.border = THIN_BORDER
    total_nca_row = bs_row
    tracker.record("Balance Sheet", "total_nca", bs_row)
    bs_row += 2  # blank row

    # --- CURRENT ASSETS ---
    ws_bs.cell(row=bs_row, column=1, value="CURRENT ASSETS").font = HEADER_FONT
    bs_row += 1

    current_asset_codes = [
        ("1100", "Trade Receivables"),
        ("1220", "Deposits"),
        ("1000", "Cash at Bank"),
        ("1010", "Cash in Hand"),
    ]
    ca_first_row = bs_row
    for acct_code, label in current_asset_codes:
        ws_bs.cell(row=bs_row, column=1, value=label).font = NORMAL_FONT
        if tracker.has("Trial Balance", acct_code):
            tb_r = tracker.row("Trial Balance", acct_code)
            # Current assets are DR balances — reference TB DR column (C)
            ca_formula = f"='Trial Balance'!C{tb_r}"
            ca_cell = ws_bs.cell(row=bs_row, column=2, value=ca_formula)
        else:
            ca_cell = ws_bs.cell(row=bs_row, column=2, value=0)
        ca_cell.number_format = CURRENCY_FMT
        ca_cell.font = NORMAL_FONT
        bs_row += 1
    ca_last_row = bs_row - 1

    # Total Current Assets
    ws_bs.cell(row=bs_row, column=1, value="Total Current Assets").font = HEADER_FONT
    total_ca_cell = ws_bs.cell(row=bs_row, column=2, value=f"=SUM(B{ca_first_row}:B{ca_last_row})")
    total_ca_cell.number_format = CURRENCY_FMT
    total_ca_cell.font = HEADER_FONT
    total_ca_cell.border = THIN_BORDER
    total_ca_row = bs_row
    tracker.record("Balance Sheet", "total_ca", bs_row)
    bs_row += 2  # blank row

    # TOTAL ASSETS
    ws_bs.cell(row=bs_row, column=1, value="TOTAL ASSETS").font = Font(bold=True, size=12, name="Arial")
    total_assets_cell = ws_bs.cell(row=bs_row, column=2, value=f"=B{total_nca_row}+B{total_ca_row}")
    total_assets_cell.number_format = CURRENCY_FMT
    total_assets_cell.font = Font(bold=True, size=12, name="Arial")
    total_assets_cell.border = DOUBLE_BORDER
    total_assets_row = bs_row
    tracker.record("Balance Sheet", "total_assets", bs_row)
    bs_row += 2  # blank row

    # --- EQUITY ---
    ws_bs.cell(row=bs_row, column=1, value="EQUITY").font = HEADER_FONT
    bs_row += 1

    # Share Capital
    ws_bs.cell(row=bs_row, column=1, value="Share Capital").font = NORMAL_FONT
    if tracker.has("Trial Balance", "3000"):
        tb_r = tracker.row("Trial Balance", "3000")
        share_cap_formula = f"='Trial Balance'!D{tb_r}"
        share_cap_cell = ws_bs.cell(row=bs_row, column=2, value=share_cap_formula)
    else:
        share_cap_cell = ws_bs.cell(row=bs_row, column=2, value=float(share_cap))
    share_cap_cell.number_format = CURRENCY_FMT
    share_cap_cell.font = NORMAL_FONT
    share_cap_bs_row = bs_row
    bs_row += 1

    # Accumulated Loss B/F — audited figure, written as value
    ws_bs.cell(row=bs_row, column=1, value="Accumulated Loss B/F").font = NORMAL_FONT
    acc_loss_cell = ws_bs.cell(row=bs_row, column=2, value=-58547.35)
    acc_loss_cell.number_format = CURRENCY_FMT
    acc_loss_cell.font = NORMAL_FONT
    acc_loss_row = bs_row
    bs_row += 1

    # Current Year Profit/(Loss) — reference IS net profit
    ws_bs.cell(row=bs_row, column=1, value="Current Year Profit / (Loss)").font = NORMAL_FONT
    np_is_row = tracker.row("Income Statement", "net_profit")
    cur_yr_formula = f"='Income Statement'!C{np_is_row}"
    cur_yr_cell = ws_bs.cell(row=bs_row, column=2, value=cur_yr_formula)
    cur_yr_cell.number_format = CURRENCY_FMT
    cur_yr_cell.font = NORMAL_FONT
    cur_yr_row = bs_row
    bs_row += 1

    # Total Equity
    ws_bs.cell(row=bs_row, column=1, value="Total Equity").font = HEADER_FONT
    total_eq_formula = f"=B{share_cap_bs_row}+B{acc_loss_row}+B{cur_yr_row}"
    total_eq_cell = ws_bs.cell(row=bs_row, column=2, value=total_eq_formula)
    total_eq_cell.number_format = CURRENCY_FMT
    total_eq_cell.font = HEADER_FONT
    total_eq_cell.border = THIN_BORDER
    total_eq_bs_row = bs_row
    tracker.record("Balance Sheet", "total_equity", bs_row)
    bs_row += 2  # blank row

    # --- CURRENT LIABILITIES ---
    ws_bs.cell(row=bs_row, column=1, value="CURRENT LIABILITIES").font = HEADER_FONT
    bs_row += 1

    liability_codes = [
        ("2000", "Trade Payables"),
        ("2100", "Other Payables & Accruals"),
        ("2350", "Amount Due to Director"),
        ("2360", "Amount Due to Shareholders"),
    ]
    cl_first_row = bs_row
    for acct_code, label in liability_codes:
        ws_bs.cell(row=bs_row, column=1, value=label).font = NORMAL_FONT
        if tracker.has("Trial Balance", acct_code):
            tb_r = tracker.row("Trial Balance", acct_code)
            # Liabilities are CR balances — reference TB CR column (D)
            cl_formula = f"='Trial Balance'!D{tb_r}"
            cl_cell = ws_bs.cell(row=bs_row, column=2, value=cl_formula)
        else:
            cl_cell = ws_bs.cell(row=bs_row, column=2, value=0)
        cl_cell.number_format = CURRENCY_FMT
        cl_cell.font = NORMAL_FONT
        bs_row += 1
    cl_last_row = bs_row - 1

    # Total Current Liabilities
    ws_bs.cell(row=bs_row, column=1, value="Total Current Liabilities").font = HEADER_FONT
    total_cl_cell = ws_bs.cell(row=bs_row, column=2, value=f"=SUM(B{cl_first_row}:B{cl_last_row})")
    total_cl_cell.number_format = CURRENCY_FMT
    total_cl_cell.font = HEADER_FONT
    total_cl_cell.border = THIN_BORDER
    total_cl_bs_row = bs_row
    tracker.record("Balance Sheet", "total_cl", bs_row)
    bs_row += 2  # blank row

    # TOTAL EQUITY AND LIABILITIES
    ws_bs.cell(row=bs_row, column=1, value="TOTAL EQUITY AND LIABILITIES").font = Font(bold=True, size=12, name="Arial")
    total_eq_liab_formula = f"=B{total_eq_bs_row}+B{total_cl_bs_row}"
    total_eq_liab_cell = ws_bs.cell(row=bs_row, column=2, value=total_eq_liab_formula)
    total_eq_liab_cell.number_format = CURRENCY_FMT
    total_eq_liab_cell.font = Font(bold=True, size=12, name="Arial")
    total_eq_liab_cell.border = DOUBLE_BORDER
    total_eq_liab_bs_row = bs_row
    bs_row += 2  # blank row

    # CHECK row
    ws_bs.cell(row=bs_row, column=1, value="CHECK (Assets - Equity & Liabilities)").font = HEADER_FONT
    bs_check_formula = f"=B{total_assets_row}-B{total_eq_liab_bs_row}"
    bs_check_cell = ws_bs.cell(row=bs_row, column=2, value=bs_check_formula)
    bs_check_cell.number_format = CURRENCY_FMT
    bs_check_cell.font = HEADER_FONT
    tracker.record("Balance Sheet", "check", bs_row)

    print(f"    NCA: 2 items, CA: {len(current_asset_codes)} items")
    print(f"    Equity: 3 items, CL: {len(liability_codes)} items")
    print(f"    BS CHECK formula at row {bs_row}")

    # ===================================================================
    # SHEET 9: FIXED ASSET REGISTER
    # ===================================================================
    print("  Sheet 9: Fixed Asset Register...")

    ws_far.column_dimensions["A"].width = 25
    ws_far.column_dimensions["B"].width = 18
    ws_far.column_dimensions["C"].width = 18
    ws_far.column_dimensions["D"].width = 18

    ws_far.cell(row=1, column=1, value="[CLIENT_NAME]").font = TITLE_FONT
    ws_far.cell(row=2, column=1, value="FIXED ASSET REGISTER").font = SUBTITLE_FONT
    ws_far.cell(row=3, column=1, value="AS AT 31 JANUARY 2025").font = SUBTITLE_FONT

    # Headers
    ws_far.cell(row=5, column=2, value="Music System &").font = HEADER_FONT
    ws_far.cell(row=5, column=3, value="Computer &").font = HEADER_FONT
    ws_far.cell(row=5, column=4, value="Total").font = HEADER_FONT
    ws_far.cell(row=6, column=2, value="Equipment").font = HEADER_FONT
    ws_far.cell(row=6, column=3, value="POS").font = HEADER_FONT

    ws_far.cell(row=7, column=1, value="Depreciation Rate").font = NORMAL_FONT
    dep_rate_b = ws_far.cell(row=7, column=2, value=0.20)
    dep_rate_b.number_format = '0%'
    dep_rate_b.font = NORMAL_FONT
    dep_rate_c = ws_far.cell(row=7, column=3, value=0.20)
    dep_rate_c.number_format = '0%'
    dep_rate_c.font = NORMAL_FONT

    # --- COST ---
    ws_far.cell(row=9, column=1, value="COST").font = HEADER_FONT

    ws_far.cell(row=10, column=1, value="Balance b/f").font = NORMAL_FONT
    ws_far.cell(row=10, column=2, value=60353.92).number_format = CURRENCY_FMT
    ws_far.cell(row=10, column=3, value=3568.00).number_format = CURRENCY_FMT
    ws_far.cell(row=10, column=4, value="=B10+C10").number_format = CURRENCY_FMT

    ws_far.cell(row=11, column=1, value="Additions").font = NORMAL_FONT
    ws_far.cell(row=11, column=2, value=0).number_format = CURRENCY_FMT
    ws_far.cell(row=11, column=3, value=0).number_format = CURRENCY_FMT
    ws_far.cell(row=11, column=4, value="=B11+C11").number_format = CURRENCY_FMT

    ws_far.cell(row=12, column=1, value="Disposals").font = NORMAL_FONT
    ws_far.cell(row=12, column=2, value=0).number_format = CURRENCY_FMT
    ws_far.cell(row=12, column=3, value=0).number_format = CURRENCY_FMT
    ws_far.cell(row=12, column=4, value="=B12+C12").number_format = CURRENCY_FMT

    ws_far.cell(row=13, column=1, value="Balance c/f").font = HEADER_FONT
    ws_far.cell(row=13, column=2, value="=B10+B11-B12").number_format = CURRENCY_FMT
    ws_far.cell(row=13, column=3, value="=C10+C11-C12").number_format = CURRENCY_FMT
    ws_far.cell(row=13, column=4, value="=D10+D11-D12").number_format = CURRENCY_FMT
    for c in range(2, 5):
        ws_far.cell(row=13, column=c).border = THIN_BORDER

    # --- ACCUMULATED DEPRECIATION ---
    ws_far.cell(row=15, column=1, value="ACCUMULATED DEPRECIATION").font = HEADER_FONT

    ws_far.cell(row=16, column=1, value="Balance b/f").font = NORMAL_FONT
    ws_far.cell(row=16, column=2, value=12070.78).number_format = CURRENCY_FMT
    ws_far.cell(row=16, column=3, value=713.60).number_format = CURRENCY_FMT
    ws_far.cell(row=16, column=4, value="=B16+C16").number_format = CURRENCY_FMT

    ws_far.cell(row=17, column=1, value="Charge for year").font = NORMAL_FONT
    ws_far.cell(row=17, column=2, value="=B13*0.2").number_format = CURRENCY_FMT
    ws_far.cell(row=17, column=3, value="=C13*0.2").number_format = CURRENCY_FMT
    ws_far.cell(row=17, column=4, value="=B17+C17").number_format = CURRENCY_FMT

    ws_far.cell(row=18, column=1, value="Balance c/f").font = HEADER_FONT
    ws_far.cell(row=18, column=2, value="=B16+B17").number_format = CURRENCY_FMT
    ws_far.cell(row=18, column=3, value="=C16+C17").number_format = CURRENCY_FMT
    ws_far.cell(row=18, column=4, value="=D16+D17").number_format = CURRENCY_FMT
    for c in range(2, 5):
        ws_far.cell(row=18, column=c).border = THIN_BORDER

    # --- NET BOOK VALUE ---
    ws_far.cell(row=20, column=1, value="NET BOOK VALUE").font = HEADER_FONT

    ws_far.cell(row=21, column=1, value="31 January 2025").font = HEADER_FONT
    ws_far.cell(row=21, column=2, value="=B13-B18").number_format = CURRENCY_FMT
    ws_far.cell(row=21, column=3, value="=C13-C18").number_format = CURRENCY_FMT
    ws_far.cell(row=21, column=4, value="=D13-D18").number_format = CURRENCY_FMT
    for c in range(2, 5):
        ws_far.cell(row=21, column=c).border = DOUBLE_BORDER
        ws_far.cell(row=21, column=c).font = HEADER_FONT

    ws_far.cell(row=22, column=1, value="31 January 2024").font = NORMAL_FONT
    ws_far.cell(row=22, column=2, value="=B10-B16").number_format = CURRENCY_FMT
    ws_far.cell(row=22, column=3, value="=C10-C16").number_format = CURRENCY_FMT
    ws_far.cell(row=22, column=4, value="=D10-D16").number_format = CURRENCY_FMT

    # Record NBV cells
    tracker.record("Fixed Assets", "nbv_music", 21)
    tracker.record("Fixed Assets", "nbv_comp", 21)

    # --- Update Balance Sheet NCA cells with FAR references ---
    nca_music_row = tracker.row("Balance Sheet", "nca_music")
    ws_bs[f"B{nca_music_row}"] = "='Fixed Assets'!B21"
    ws_bs[f"B{nca_music_row}"].number_format = CURRENCY_FMT
    ws_bs[f"B{nca_music_row}"].font = NORMAL_FONT

    nca_comp_row = tracker.row("Balance Sheet", "nca_comp")
    ws_bs[f"B{nca_comp_row}"] = "='Fixed Assets'!C21"
    ws_bs[f"B{nca_comp_row}"].number_format = CURRENCY_FMT
    ws_bs[f"B{nca_comp_row}"].font = NORMAL_FONT

    print(f"    FAR written with formula-driven NBV")
    print(f"    BS NCA cells (rows {nca_music_row}, {nca_comp_row}) updated with FAR references")

    # ===================================================================
    # SHEET 10: PAYROLL SUMMARY
    # ===================================================================
    print("  Sheet 10: Payroll Summary...")

    ws_payroll.column_dimensions["A"].width = 12
    ws_payroll.column_dimensions["B"].width = 50
    ws_payroll.column_dimensions["C"].width = 16

    ws_payroll.cell(row=1, column=1, value="[CLIENT_NAME]").font = TITLE_FONT
    ws_payroll.cell(row=2, column=1, value="PAYROLL SUMMARY").font = SUBTITLE_FONT
    ws_payroll.cell(row=3, column=1, value="FY 1 FEBRUARY 2024 TO 31 JANUARY 2025").font = SUBTITLE_FONT

    # Headers
    pay_headers = ["Date", "Employee/Description", "Amount (RM)"]
    for c, h in enumerate(pay_headers, start=1):
        cell = ws_payroll.cell(row=5, column=c, value=h)
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    # Extract salary transactions (acct_code "5000" AFTER reclassification)
    salary_txns = [txn for txn in all_txns if txn["acct_code"] == "5000"]

    # Group by month (YYYY-MM)
    from collections import OrderedDict as OD
    salary_by_month: dict[str, list] = OrderedDict()
    for txn in salary_txns:
        month_key = txn["date"][:7]  # "2024-02"
        if month_key not in salary_by_month:
            salary_by_month[month_key] = []
        salary_by_month[month_key].append(txn)

    def extract_payee(desc: str) -> str:
        """Extract payee name from description like 'TRANSFER FR A/C | NAME | ...'"""
        parts = desc.split("|")
        if len(parts) >= 2:
            name = parts[1].strip()
            # Clean up trailing asterisks and whitespace
            name = name.rstrip("* ").strip()
            return name[:50]
        return desc[:50]

    MONTH_NAMES = {
        "01": "January", "02": "February", "03": "March", "04": "April",
        "05": "May", "06": "June", "07": "July", "08": "August",
        "09": "September", "10": "October", "11": "November", "12": "December",
    }

    pay_row = 6
    subtotal_cells = []

    for month_key, txns in salary_by_month.items():
        year_str, month_num = month_key.split("-")
        month_name = f"{MONTH_NAMES[month_num]} {year_str}"

        # Month header
        ws_payroll.cell(row=pay_row, column=1, value=month_name).font = HEADER_FONT
        pay_row += 1

        month_first_row = pay_row
        for txn in txns:
            ws_payroll.cell(row=pay_row, column=1, value=txn["date"]).font = NORMAL_FONT
            payee = extract_payee(txn["description"])
            ws_payroll.cell(row=pay_row, column=2, value=payee).font = NORMAL_FONT
            amt_cell = ws_payroll.cell(row=pay_row, column=3, value=abs(float(d(txn["amount"]))))
            amt_cell.number_format = CURRENCY_FMT
            amt_cell.font = NORMAL_FONT
            pay_row += 1
        month_last_row = pay_row - 1

        # Subtotal
        ws_payroll.cell(row=pay_row, column=2, value="Subtotal").font = HEADER_FONT
        subtotal_formula = f"=SUM(C{month_first_row}:C{month_last_row})"
        sub_cell = ws_payroll.cell(row=pay_row, column=3, value=subtotal_formula)
        sub_cell.number_format = CURRENCY_FMT
        sub_cell.font = HEADER_FONT
        sub_cell.border = THIN_BORDER
        subtotal_cells.append(f"C{pay_row}")
        pay_row += 2  # blank row after subtotal

    # Grand Total
    ws_payroll.cell(row=pay_row, column=2, value="GRAND TOTAL").font = Font(bold=True, size=12, name="Arial")
    grand_total_formula = "=" + "+".join(subtotal_cells)
    gt_cell = ws_payroll.cell(row=pay_row, column=3, value=grand_total_formula)
    gt_cell.number_format = CURRENCY_FMT
    gt_cell.font = Font(bold=True, size=12, name="Arial")
    gt_cell.border = DOUBLE_BORDER
    grand_total_row = pay_row
    pay_row += 1

    # CHECK vs TB
    if tracker.has("Trial Balance", "5000"):
        salary_tb_row = tracker.row("Trial Balance", "5000")
        ws_payroll.cell(row=pay_row, column=2, value="CHECK vs Trial Balance").font = HEADER_FONT
        check_formula = f"=C{grand_total_row}-'Trial Balance'!C{salary_tb_row}"
        check_cell = ws_payroll.cell(row=pay_row, column=3, value=check_formula)
        check_cell.number_format = CURRENCY_FMT
        check_cell.font = HEADER_FONT
        pay_row += 1

    # Note
    pay_row += 1
    ws_payroll.cell(row=pay_row, column=1, value="Note: EPF/SOCSO/EIS status pending client confirmation — see Queries & Notes #4").font = NORMAL_FONT

    print(f"    {len(salary_txns)} salary transactions across {len(salary_by_month)} months")
    print(f"    Grand total at row {grand_total_row}")

    # ===================================================================
    # SHEET 11: TAX COMPUTATION
    # ===================================================================
    print("  Sheet 11: Tax Computation...")

    ws_tax.column_dimensions["A"].width = 35
    ws_tax.column_dimensions["B"].width = 14
    ws_tax.column_dimensions["C"].width = 14
    ws_tax.column_dimensions["D"].width = 10
    ws_tax.column_dimensions["E"].width = 14
    ws_tax.column_dimensions["F"].width = 14

    ws_tax.cell(row=1, column=1, value="[CLIENT_NAME]").font = TITLE_FONT
    ws_tax.cell(row=2, column=1, value="TAX COMPUTATION \u2014 YA 2025").font = SUBTITLE_FONT
    ws_tax.cell(row=3, column=1, value="BASIS PERIOD: 1 FEBRUARY 2024 TO 31 JANUARY 2025").font = SUBTITLE_FONT

    # --- SECTION A: BUSINESS INCOME SCHEDULE ---
    ws_tax.cell(row=5, column=1, value="BUSINESS INCOME SCHEDULE").font = HEADER_FONT

    # Net Loss per Accounts
    net_profit_is_row = tracker.row("Income Statement", "net_profit")
    ws_tax.cell(row=7, column=1, value="Net Loss per Accounts").font = NORMAL_FONT
    np_ref = ws_tax.cell(row=7, column=3, value=f"='Income Statement'!C{net_profit_is_row}")
    np_ref.number_format = CURRENCY_FMT
    np_ref.font = NORMAL_FONT

    # Add: Non-allowable expenses
    ws_tax.cell(row=9, column=1, value="Add: Non-allowable expenses").font = HEADER_FONT

    # Depreciation of PPE
    depr_music_row = tracker.row("Income Statement", "depr_music")
    depr_comp_row = tracker.row("Income Statement", "depr_comp")
    ws_tax.cell(row=10, column=1, value="Depreciation of PPE").font = NORMAL_FONT
    depr_formula = f"=ABS('Income Statement'!C{depr_music_row})+ABS('Income Statement'!C{depr_comp_row})"
    depr_cell = ws_tax.cell(row=10, column=2, value=depr_formula)
    depr_cell.number_format = CURRENCY_FMT
    depr_cell.font = NORMAL_FONT

    # Auditors' remuneration
    ws_tax.cell(row=11, column=1, value="Auditors' remuneration").font = NORMAL_FONT
    ws_tax.cell(row=11, column=2, value=3000.00).number_format = CURRENCY_FMT

    # Accounting fee
    ws_tax.cell(row=12, column=1, value="Accounting fee").font = NORMAL_FONT
    ws_tax.cell(row=12, column=2, value=2000.00).number_format = CURRENCY_FMT

    # Total add-backs
    ws_tax.cell(row=13, column=2, value="Total add-backs").font = HEADER_FONT
    add_back_cell = ws_tax.cell(row=13, column=3, value="=SUM(B10:B12)")
    add_back_cell.number_format = CURRENCY_FMT
    add_back_cell.font = HEADER_FONT

    # Less: Special deductions
    ws_tax.cell(row=15, column=1, value="Less: Special deductions (S34(6))").font = HEADER_FONT

    ws_tax.cell(row=16, column=1, value="Audit expenditure S34(6)(e)").font = NORMAL_FONT
    ws_tax.cell(row=16, column=2, value=-3000.00).number_format = CURRENCY_FMT

    ws_tax.cell(row=17, column=1, value="Accounting fee S34(6)(e)").font = NORMAL_FONT
    ws_tax.cell(row=17, column=2, value=-2000.00).number_format = CURRENCY_FMT

    ws_tax.cell(row=18, column=2, value="Total deductions").font = HEADER_FONT
    ded_cell = ws_tax.cell(row=18, column=3, value="=SUM(B16:B17)")
    ded_cell.number_format = CURRENCY_FMT
    ded_cell.font = HEADER_FONT

    # Net adjustment
    ws_tax.cell(row=20, column=2, value="Net adjustment").font = NORMAL_FONT
    net_adj_cell = ws_tax.cell(row=20, column=3, value="=C13+C18")
    net_adj_cell.number_format = CURRENCY_FMT
    net_adj_cell.font = NORMAL_FONT

    # Adjusted Business Income/(Loss)
    ws_tax.cell(row=21, column=1, value="ADJUSTED BUSINESS INCOME/(LOSS)").font = HEADER_FONT
    adj_inc_cell = ws_tax.cell(row=21, column=3, value="=C7+C20")
    adj_inc_cell.number_format = CURRENCY_FMT
    adj_inc_cell.font = HEADER_FONT

    # Less: Capital allowance absorbed (references row 37 Total available)
    ws_tax.cell(row=22, column=1, value="Less: Capital allowance absorbed").font = NORMAL_FONT
    ca_absorbed_cell = ws_tax.cell(row=22, column=3, value="=-MIN(C37,MAX(0,C21))")
    ca_absorbed_cell.number_format = CURRENCY_FMT
    ca_absorbed_cell.font = NORMAL_FONT

    # Statutory Business Income
    ws_tax.cell(row=23, column=1, value="STATUTORY BUSINESS INCOME").font = HEADER_FONT
    stat_inc_cell = ws_tax.cell(row=23, column=3, value="=MAX(0,C21-MIN(C37,MAX(0,C21)))")
    stat_inc_cell.number_format = CURRENCY_FMT
    stat_inc_cell.font = HEADER_FONT
    stat_inc_cell.border = THIN_BORDER

    # --- SECTION B: CAPITAL ALLOWANCE SCHEDULE ---
    ws_tax.cell(row=25, column=1, value="CAPITAL ALLOWANCE SCHEDULE").font = HEADER_FONT

    ca_headers = ["Asset", "Original Cost", "RE b/f", "AA Rate", "Annual Allowance", "RE c/f"]
    for c, h in enumerate(ca_headers, start=1):
        cell = ws_tax.cell(row=26, column=c, value=h)
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    # CA schedule data: (asset, cost, re_bf, rate_display, has_formula)
    ca_assets = [
        ("POS Receipt Printer",       3568,  2497,  0.10, True),
        ("Audio & Lighting (Bose)",   44052, 30837, 0.10, True),
        ("Guitar Amplifier",           1992,  0,    None,  False),
        ("PDP Drumset",                5500,  3850,  0.10, True),
        ("Korg Granstage Piano",       8810,  6167,  0.10, True),
    ]

    ca_start_row = 27
    for i, (asset, cost, re_bf, rate, has_formula) in enumerate(ca_assets):
        r = ca_start_row + i
        ws_tax.cell(row=r, column=1, value=asset).font = NORMAL_FONT
        ws_tax.cell(row=r, column=2, value=float(cost)).number_format = CURRENCY_FMT
        ws_tax.cell(row=r, column=3, value=float(re_bf)).number_format = CURRENCY_FMT

        if has_formula:
            rate_cell = ws_tax.cell(row=r, column=4, value=rate)
            rate_cell.number_format = '0%'
            rate_cell.font = NORMAL_FONT
            ws_tax.cell(row=r, column=5, value=f"=B{r}*0.1").number_format = CURRENCY_FMT
            ws_tax.cell(row=r, column=6, value=f"=C{r}-E{r}").number_format = CURRENCY_FMT
        else:
            ws_tax.cell(row=r, column=4, value="-").font = NORMAL_FONT
            ws_tax.cell(row=r, column=5, value=0).number_format = CURRENCY_FMT
            ws_tax.cell(row=r, column=6, value=0).number_format = CURRENCY_FMT

    ca_end_row = ca_start_row + len(ca_assets) - 1  # row 31

    # Total row
    ca_total_row = ca_end_row + 1  # row 32
    ws_tax.cell(row=ca_total_row, column=1, value="TOTAL").font = HEADER_FONT
    ws_tax.cell(row=ca_total_row, column=2, value=f"=SUM(B{ca_start_row}:B{ca_end_row})").number_format = CURRENCY_FMT
    ws_tax.cell(row=ca_total_row, column=3, value=f"=SUM(C{ca_start_row}:C{ca_end_row})").number_format = CURRENCY_FMT
    ws_tax.cell(row=ca_total_row, column=4, value="").font = NORMAL_FONT
    ws_tax.cell(row=ca_total_row, column=5, value=f"=SUM(E{ca_start_row}:E{ca_end_row})").number_format = CURRENCY_FMT
    ws_tax.cell(row=ca_total_row, column=6, value=f"=SUM(F{ca_start_row}:F{ca_end_row})").number_format = CURRENCY_FMT
    for c in range(2, 7):
        ws_tax.cell(row=ca_total_row, column=c).font = HEADER_FONT
        ws_tax.cell(row=ca_total_row, column=c).border = THIN_BORDER

    # --- UNABSORBED CA POOL ---
    ws_tax.cell(row=34, column=1, value="UNABSORBED CAPITAL ALLOWANCE").font = HEADER_FONT

    ws_tax.cell(row=35, column=1, value="Balance b/f (from YA 2024)").font = NORMAL_FONT
    ws_tax.cell(row=35, column=3, value=20571.0).number_format = CURRENCY_FMT

    ws_tax.cell(row=36, column=1, value="Current year allowance").font = NORMAL_FONT
    ws_tax.cell(row=36, column=3, value=f"=E{ca_total_row}").number_format = CURRENCY_FMT

    ws_tax.cell(row=37, column=1, value="Total available").font = NORMAL_FONT
    ws_tax.cell(row=37, column=3, value="=C35+C36").number_format = CURRENCY_FMT
    tracker.record("Tax Computation", "ca_available", 37)

    ws_tax.cell(row=38, column=1, value="Less: Absorbed").font = NORMAL_FONT
    ws_tax.cell(row=38, column=3, value="=-MIN(C37,MAX(0,C21))").number_format = CURRENCY_FMT

    ws_tax.cell(row=39, column=1, value="Balance c/f").font = HEADER_FONT
    cf_cell = ws_tax.cell(row=39, column=3, value="=C37+C38")
    cf_cell.number_format = CURRENCY_FMT
    cf_cell.font = HEADER_FONT
    cf_cell.border = THIN_BORDER

    # --- SECTION C: LOSS UTILIZATION ---
    ws_tax.cell(row=41, column=1, value="UNABSORBED BUSINESS LOSSES").font = HEADER_FONT

    ws_tax.cell(row=42, column=1, value="Losses b/f (from YA 2024)").font = NORMAL_FONT
    ws_tax.cell(row=42, column=3, value=36978.0).number_format = CURRENCY_FMT

    ws_tax.cell(row=43, column=1, value="Current year adjusted loss").font = NORMAL_FONT
    ws_tax.cell(row=43, column=3, value="=IF(C21<0,ABS(C21),0)").number_format = CURRENCY_FMT

    ws_tax.cell(row=44, column=1, value="Total losses").font = NORMAL_FONT
    ws_tax.cell(row=44, column=3, value="=C42+C43").number_format = CURRENCY_FMT

    ws_tax.cell(row=45, column=1, value="Less: Absorbed").font = NORMAL_FONT
    ws_tax.cell(row=45, column=3, value=0).number_format = CURRENCY_FMT

    ws_tax.cell(row=46, column=1, value="Losses c/f").font = HEADER_FONT
    losses_cf_cell = ws_tax.cell(row=46, column=3, value="=C44-C45")
    losses_cf_cell.number_format = CURRENCY_FMT
    losses_cf_cell.font = HEADER_FONT
    losses_cf_cell.border = THIN_BORDER

    ws_tax.cell(row=47, column=1, value="Note: 10-year carry-forward limit. YA 2024 loss expires YA 2034.").font = NORMAL_FONT

    # --- SECTION D: TAX PAYABLE ---
    ws_tax.cell(row=49, column=1, value="TAX PAYABLE").font = HEADER_FONT

    ws_tax.cell(row=50, column=1, value="Chargeable income").font = NORMAL_FONT
    ws_tax.cell(row=50, column=3, value="=MAX(0,C23)").number_format = CURRENCY_FMT

    ws_tax.cell(row=51, column=1, value="Tax at 15% (first RM150,000)").font = NORMAL_FONT
    ws_tax.cell(row=51, column=3, value="=MIN(C50,150000)*0.15").number_format = CURRENCY_FMT

    ws_tax.cell(row=52, column=1, value="Tax at 24% (remainder)").font = NORMAL_FONT
    ws_tax.cell(row=52, column=3, value="=MAX(0,C50-150000)*0.24").number_format = CURRENCY_FMT

    ws_tax.cell(row=53, column=1, value="TOTAL TAX PAYABLE").font = HEADER_FONT
    tax_total_cell = ws_tax.cell(row=53, column=3, value="=C51+C52")
    tax_total_cell.number_format = CURRENCY_FMT
    tax_total_cell.font = HEADER_FONT
    tax_total_cell.border = DOUBLE_BORDER

    print(f"    Tax computation written with CA schedule ({len(ca_assets)} assets)")
    print(f"    CA Total row: {ca_total_row}, CA available row: 37")

    # ===================================================================
    # SHEET 12: QUERIES & NOTES
    # ===================================================================
    print("  Sheet 12: Queries & Notes...")

    ws_queries.column_dimensions["A"].width = 5
    ws_queries.column_dimensions["B"].width = 15
    ws_queries.column_dimensions["C"].width = 30
    ws_queries.column_dimensions["D"].width = 80
    ws_queries.column_dimensions["E"].width = 10

    ws_queries.cell(row=1, column=1, value="[CLIENT_NAME]").font = TITLE_FONT
    ws_queries.cell(row=2, column=1, value="QUERIES & NOTES").font = SUBTITLE_FONT
    ws_queries.cell(row=3, column=1, value="FY 1 FEBRUARY 2024 TO 31 JANUARY 2025").font = SUBTITLE_FONT

    q_headers = ["#", "Category", "Item", "Detail", "Status"]
    for c, h in enumerate(q_headers, start=1):
        cell = ws_queries.cell(row=5, column=c, value=h)
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    queries_data = [
        (1, "Assumption", "Prior year accrual settlements",
         "First Feb 2024 salary payments (Riduan RM3,945 + Aqmar RM2,370 + Vandolf RM1,900 = RM8,215) "
         "reclassified as Jan 2024 wages accrual. First TNB payment (RM3,000) reclassified as electricity accrual. "
         "Total RM11,215 reclassified. Remaining RM4,500 (audit RM3,000 + tax agent RM1,500) stays as payable.",
         "Assumed"),

        (2, "Query", "Trade Receivables RM17,907",
         "Liberal Latte Coffee — outstanding 12+ months with zero movement. "
         "Impairment assessment required under MPERS S11. Please confirm collectability or approve provision/write-off.",
         "Open"),

        (3, "Query", "Cash in Hand RM6,035",
         "Petty cash float with zero activity for 12 months. Please confirm petty cash balance by physical count.",
         "Open"),

        (4, "Query", "Statutory Contributions (EPF/SOCSO/EIS)",
         "Salary of RM93,497 recorded but no employer statutory contributions. "
         "Are employees registered for EPF/SOCSO/EIS? Does employer absorb employee contributions? "
         "Please provide payslips or confirm.",
         "Open"),

        (5, "Going Concern", "Deficit shareholders' fund",
         "Company has accumulated losses of RM71,133, deficit shareholders' fund, and net current liabilities of RM109,487. "
         "Prior year audit report flagged material uncertainty. Directors' letter of financial support required for going concern basis.",
         "Open"),

        (6, "Query", "Trade Payables RM63,810",
         "Largely unchanged from prior year (RM64,810 less RM1,000 payment). "
         "Please confirm these remain valid outstanding obligations and provide supplier statements.",
         "Open"),

        (7, "Note", "Bank Reconciliation",
         "GL closing balance RM12,513.29 vs bank statement RM12,511.57. "
         "Difference of RM1.72 due to rounding accumulation across 1,423 transactions in CSV extraction. "
         "Within materiality but documented.",
         "Noted"),

        (8, "Note", "Prior Year Tax Carry-Forward",
         "From YA 2024: Unabsorbed capital allowances RM20,571 and business losses RM36,978 brought forward "
         "per tax computation. 10-year loss carry-forward limit applies (YA 2024 loss expires YA 2034).",
         "Noted"),
    ]

    for r_offset, (num, cat, item, detail, status) in enumerate(queries_data):
        r = 6 + r_offset
        ws_queries.cell(row=r, column=1, value=num).font = NORMAL_FONT
        ws_queries.cell(row=r, column=2, value=cat).font = NORMAL_FONT
        ws_queries.cell(row=r, column=3, value=item).font = NORMAL_FONT
        ws_queries.cell(row=r, column=4, value=detail).font = NORMAL_FONT
        ws_queries.cell(row=r, column=5, value=status).font = NORMAL_FONT

    print(f"    {len(queries_data)} queries/notes written")

    # ===================================================================
    # SAVE WORKBOOK
    # ===================================================================

    wb.save(OUTPUT)
    print(f"\nWorkbook saved to: {OUTPUT}")
    print("All 12 sheets complete!")
    print("DONE!")